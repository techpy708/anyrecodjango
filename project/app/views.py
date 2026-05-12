# =============================
# Standard Library Imports
# =============================
import base64
import datetime
import gc
import io
import json
import os
import re
import tempfile
import zipfile
from collections import defaultdict
from datetime import datetime
from functools import wraps
from io import BytesIO


# Create your views here.
from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages

from django.http import HttpResponseForbidden
from django.shortcuts import render, redirect

import os
import tempfile
import pandas as pd
from datetime import datetime
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.utils.text import get_valid_filename

from werkzeug.utils import secure_filename
import pandas as pd
import json
import os
import duckdb
import re
import tempfile
import zipfile
import io
import gc
from datetime import datetime
from rapidfuzz import fuzz
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import user_passes_test
from django.contrib import messages
from django.utils import timezone
from django.db.models import Sum

import os
import json
import io
import zipfile
import tempfile
import gc
import traceback
from datetime import datetime
import pandas as pd
import duckdb
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.utils.text import get_valid_filename


# =============================
# Third-Party Library Imports
# =============================
import duckdb
import pandas as pd
import requests
from docx import Document as DocxDocument
from docx.enum.text import WD_ALIGN_PARAGRAPH
from docx.oxml import OxmlElement
from docx.oxml.ns import qn
from docx.shared import Pt, RGBColor
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz
from werkzeug.utils import secure_filename

# =============================
# Django Imports
# =============================
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import (
    get_user_model,
    login,
    update_session_auth_hash,
    views as auth_views,
)
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.hashers import make_password
from django.core.exceptions import ObjectDoesNotExist, PermissionDenied
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Count, Prefetch, Sum
from django.http import (
    FileResponse,
    Http404,
    HttpResponse,
    HttpResponseForbidden,
    JsonResponse,
)
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.dateformat import format as date_format
from django.utils.encoding import smart_str
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST

# =============================
# Local App Imports
# =============================
from .forms import CustomUserForm
from .models import (
    CompanyGroup,
    CustomUser,
    GroupLicense,
    ProcessingHistory,
)
from .utils import (
    clear_session_keys,
    generate_otp,
    generate_username_from_email,
    is_otp_expired,
    send_otp_email,
    send_user_credentials_email,
)


User = get_user_model()

# =============================
# AUTHENTICATION VIEWS
# =============================

# def signup(request):
#     """Handle user signup with OTP verification."""
#     if request.method == 'POST':
#         # --- ADDED: Get first_name and last_name from the form ---
#         first_name = request.POST.get('first_name', '').strip()
#         last_name = request.POST.get('last_name', '').strip()
#         email = request.POST.get('email', '').strip()
#         raw_password = request.POST.get('password', '').strip()

#         # --- UPDATED: Basic validation for new fields ---
#         if not all([first_name, last_name, email, raw_password]):
#             return render(request, 'signup.html', {'error': 'All fields are required.'})

#         if User.objects.filter(email__iexact=email).exists():
#             return render(request, 'signup.html', {'error': 'This email address is already registered.'})

#         otp = generate_otp()
#         print(otp)
        
#         # --- UPDATED: Store the new fields in the session ---
#         request.session.update({
#             'signup_otp': otp,
#             'signup_email': email,
#             'signup_password': raw_password,
#             'signup_first_name': first_name, # ADDED
#             'signup_last_name': last_name,   # ADDED
#             'signup_otp_time': datetime.now().isoformat()
#         })

#         if not send_otp_email(email, otp, 'Your Signup OTP', f'Your OTP for sign up is {otp}'):
#             return render(request, 'signup.html', {'error': 'Failed to send OTP. Please try again.'})

#         return redirect('signup_verify_otp')

#     return render(request, 'signup.html')


# def signup_verify_otp(request):
#     """Verify signup OTP and create user account."""
#     if request.method == 'POST':
#         user_otp = request.POST.get('otp', '').strip()
        
#         # --- UPDATED: Get all data, including new fields, from the session ---
#         session_data = request.session
#         stored_otp = session_data.get('signup_otp')
#         email = session_data.get('signup_email')
#         raw_password = session_data.get('signup_password')
#         first_name = session_data.get('signup_first_name') # ADDED
#         last_name = session_data.get('signup_last_name')   # ADDED
#         otp_time = session_data.get('signup_otp_time')
        
#         session_keys_to_clear = [
#             'signup_otp', 'signup_email', 'signup_password', 
#             'signup_first_name', 'signup_last_name', 'signup_otp_time'
#         ]

#         if not otp_time or is_otp_expired(otp_time):
#             clear_session_keys(request.session, session_keys_to_clear)
#             return render(request, 'signup_verify_otp.html', {'error': 'OTP expired. Please sign up again.'})

#         if user_otp != stored_otp:
#             return render(request, 'signup_verify_otp.html', {'error': 'Invalid OTP. Please try again.'})

#         try:
#             # --- UPDATED: Create user with first_name and last_name ---
#             user = User(
#                 email=email,
#                 first_name=first_name,
#                 last_name=last_name
#             )
#             # Your existing logic for username, password, etc.
#             if hasattr(user, 'username'):
#                 user.username = generate_username_from_email(email)
#             user.set_password(raw_password)
#             if hasattr(user, 'access_rights'):
#                 user.access_rights = 'admin'  # Or 'download' as a safer default
            
#             user.save()

#             # Log the user in
#             user.backend = 'app.backends.EmailOrUsernameModelBackend' # Ensure this path is correct
#             login(request, user)
#             clear_session_keys(request.session, session_keys_to_clear)

#             return redirect('dashboard_view')

#         except Exception as e:
#             return render(request, 'signup_verify_otp.html', {'error': f'Account creation failed. Error: {e}'})

#     return render(request, 'signup_verify_otp.html')

# --- SIGNUP VIEW ---
def signup(request):
    """Handle user signup with OTP verification."""
    if request.method == 'POST':
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        group_name = request.POST.get('group_name', '').strip() # --- NEW FIELD
        email = request.POST.get('email', '').strip()
        raw_password = request.POST.get('password', '').strip()

        # --- VALIDATION ---
        if not all([first_name, last_name, group_name, email, raw_password]):
            return render(request, 'signup.html', {'error': 'All fields, including Group Name, are required.'})

        if User.objects.filter(email__iexact=email).exists():
            return render(request, 'signup.html', {'error': 'This email address is already registered.'})

        otp = generate_otp()
        print(f"Generated OTP: {otp}")
        
        # --- STORE IN SESSION ---
        request.session.update({
            'signup_otp': otp,
            'signup_email': email,
            'signup_password': raw_password,
            'signup_first_name': first_name,
            'signup_last_name': last_name,
            'signup_group_name': group_name, # --- STORED HERE
            'signup_otp_time': datetime.now().isoformat()
        })

        if not send_otp_email(email, otp, 'Your Signup OTP', f'Your OTP for sign up is {otp}'):
            return render(request, 'signup.html', {'error': 'Failed to send OTP. Please try again.'})

        return redirect('signup_verify_otp')

    return render(request, 'signup.html')


# --- VERIFY OTP VIEW ---
def signup_verify_otp(request):
    """Verify signup OTP, create user, create group, and assign user."""
    if request.method == 'POST':
        user_otp = request.POST.get('otp', '').strip()
        
        session_data = request.session
        stored_otp = session_data.get('signup_otp')
        email = session_data.get('signup_email')
        raw_password = session_data.get('signup_password')
        first_name = session_data.get('signup_first_name')
        last_name = session_data.get('signup_last_name')
        group_name = session_data.get('signup_group_name') # --- RETRIEVE GROUP NAME
        otp_time = session_data.get('signup_otp_time')
        
        session_keys_to_clear = [
            'signup_otp', 'signup_email', 'signup_password', 
            'signup_first_name', 'signup_last_name', 'signup_group_name', 'signup_otp_time'
        ]

        # Check OTP Validity
        if not otp_time or is_otp_expired(otp_time):
            clear_session_keys(request.session, session_keys_to_clear)
            return render(request, 'signup_verify_otp.html', {'error': 'OTP expired. Please sign up again.'})

        if user_otp != stored_otp:
            return render(request, 'signup_verify_otp.html', {'error': 'Invalid OTP. Please try again.'})

        try:
            # 1. Create User
            user = User(
                email=email,
                first_name=first_name,
                last_name=last_name,
                is_active=True 
            )
            if hasattr(user, 'username'):
                user.username = generate_username_from_email(email)
            user.set_password(raw_password)
            user.save()

            # 2. Create Company Group
            # Logic: If name exists, append a suffix to ensure DB uniqueness,
            # as the Model requires name to be unique.
            final_group_name = group_name
            counter = 1
            group_created = False
            
            created_group = None
            
            while not group_created:
                try:
                    created_group = CompanyGroup.objects.create(
                        name=final_group_name,
                        created_by=user
                    )
                    group_created = True
                except:
                    # If name is taken, try "Name (1)", "Name (2)", etc.
                    final_group_name = f"{group_name} ({counter})"
                    counter += 1

            # 3. Assign User to Group & Set Role
            user.company_group = created_group
            user.access_rights = 'admin' # The creator is the Admin
            user.save()

            # 4. Login and Cleanup
            user.backend = 'django.contrib.auth.backends.ModelBackend' # Or your custom backend path
            login(request, user)
            clear_session_keys(request.session, session_keys_to_clear)

            return redirect('dashboard_view')

        except Exception as e:
            # In production, log 'e' properly
            return render(request, 'signup_verify_otp.html', {'error': f'Account creation failed. Error: {str(e)}'})

    return render(request, 'signup_verify_otp.html')


def forgot_password(request):
    """Step 1: User enters email to request password reset OTP"""
    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        if not email:
            return render(request, 'forgot_password.html', {'error': 'Email is required.'})

        if not User.objects.filter(email=email).exists():
            messages.error(request, 'No account found with this email.')
            return redirect('signup')

        # Generate OTP
        otp = generate_otp()
        request.session['otp'] = otp
        request.session['email'] = email
        request.session['otp_time'] = datetime.now().isoformat()

        email_sent = send_otp_email(
            email,
            otp,
            'Password Reset OTP',
            f'Your verification code is {otp}'
        )
        if not email_sent:
            return render(request, 'forgot_password.html', {'error': 'Failed to send OTP. Please try again.'})

        return redirect('verify_otp')
    return render(request, 'forgot_password.html')


def verify_otp(request):
    """Step 2: Verify OTP before showing password reset form"""
    if request.method == 'POST':
        user_otp = request.POST.get('otp', '').strip()
        stored_otp = request.session.get('otp')
        email = request.session.get('email')
        otp_time = request.session.get('otp_time')

        if is_otp_expired(otp_time):
            clear_session_keys(request.session, ['otp', 'email', 'otp_time'])
            return render(request, 'verify_otp.html', {'error': 'OTP expired. Please request again.'})

        if user_otp == stored_otp and email:
            # OTP correct → go to reset password page
            request.session['otp_verified'] = True
            return redirect('reset_password')
        else:
            return render(request, 'verify_otp.html', {'error': 'Invalid OTP. Please try again.'})

    return render(request, 'verify_otp.html')


def reset_password(request):
    if not request.session.get('otp_verified'):
        return redirect('forgot_password')

    email = request.session.get('email')
    if request.method == 'POST':
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')

        if not password1 or not password2:
            return render(request, 'reset_password.html', {'error': 'Both fields are required.'})

        if password1 != password2:
            return render(request, 'reset_password.html', {'error': 'Passwords do not match.'})

        try:
            user = User.objects.get(email=email)
            user.set_password(password1)
            user.save()

            # ✅ Set backend before logging in
            user.backend = 'app.backends.EmailOrUsernameModelBackend'
            login(request, user)

            clear_session_keys(request.session, ['otp', 'email', 'otp_time', 'otp_verified'])

            return redirect('dashboard_view')
        except User.DoesNotExist:
            return render(request, 'reset_password.html', {'error': 'User not found. Please sign up.'})

    return render(request, 'reset_password.html')


class CustomLoginView(auth_views.LoginView):
    template_name = 'login.html'
    redirect_authenticated_user = True
    next_page = 'dashboard_view'

    def form_valid(self, form):
        remember_me = self.request.POST.get('remember_me')
        if remember_me == 'on':
            self.request.session.set_expiry(2592000)  # 30 days
        else:
            self.request.session.set_expiry(0)       # Browser close
        return super().form_valid(form)


def admin_or_superuser_required(view_func):
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if not (request.user.is_superuser or request.user.access_rights == 'admin'):
            raise PermissionDenied
        return view_func(request, *args, **kwargs)
    return _wrapped_view


@login_required
def dashboard(request):
    return render(request, 'reco.html')

@login_required
def dashboard_view(request):
    return render(request, 'reco.html')


# =============================
# USER MANAGEMENT VIEWS
# =============================

# @login_required
# def add_user(request):
#     """
#     Handles user creation and lists all manageable users.
#     Provides data for the form and the dynamic dropdowns.
#     """
#     user = request.user
#     if not (user.is_superuser or user.access_rights == 'admin'):
#         messages.error(request, "You do not have permission to access this page.")
#         return redirect('dashboard_view') 

#     if request.method == 'POST':
#         form = CustomUserForm(request.POST, request_user=request.user)
#         if form.is_valid():
#             new_user = form.save(commit=False)
#             new_user.set_password(form.cleaned_data['password2'])
#             new_user.save()
#             form.save_m2m() # Save M2M relationships

#             raw_password = form.cleaned_data.get('password2')
#             email_sent = send_user_credentials_email(new_user, raw_password)
            
#             if email_sent:
#                 messages.success(request, f"User '{new_user.username}' was created and a welcome email has been sent.")
#             else:
#                 messages.warning(request, f"User '{new_user.username}' was created, but the welcome email could not be sent.")
            
#             return redirect('add_user')
#         else:
#             messages.error(request, "Please correct the errors below and try again.")
#     else:
#         form = CustomUserForm(request_user=user)

#     # --- Query users and groups based on permissions ---
#     if user.is_superuser:
#         users = CustomUser.objects.prefetch_related('groups', 'companies', 'departments').all().order_by('username')
#         # MODIFIED: Use CompanyGroup and correct the prefetch related_name
#         groups = CompanyGroup.objects.prefetch_related('companies__departments').all()
#     elif user.access_rights == 'admin':
#         admin_groups = user.groups.all()
#         # MODIFIED: Use CompanyGroup and correct the prefetch related_name
#         groups = CompanyGroup.objects.filter(id__in=admin_groups.values_list('id', flat=True)).prefetch_related('companies__departments')
#         users = CustomUser.objects.filter(groups__in=admin_groups).prefetch_related('groups', 'companies', 'departments').distinct().order_by('username')
#     else:
#         users = CustomUser.objects.none()
#         # MODIFIED: Use CompanyGroup
#         groups = CompanyGroup.objects.none()

#     # --- Prepare JSON Data for Frontend Dropdowns (Efficiently) ---
#     groups_to_companies = {group.id: [] for group in groups}
#     companies_to_departments = {}
    
#     # MODIFIED: Query Company model based on the correct group field.
#     # The 'group' field in Company is a ForeignKey to CompanyGroup.
#     all_companies_qs = Company.objects.filter(group__in=groups).prefetch_related('departments')

#     for company in all_companies_qs:
#         if company.group_id in groups_to_companies:
#             groups_to_companies[company.group_id].append({'id': company.id, 'name': company.name})
        
#         departments_in_company = [{'id': d.id, 'name': d.name} for d in company.departments.all()]
#         companies_to_departments[company.id] = departments_in_company
    
#     all_groups_for_frontend = list(groups.values('id', 'name'))

#     context = {
#         'form': form,
#         'users': users,
#         'groups': groups,
#         'all_groups_json': json.dumps(all_groups_for_frontend),
#         'groups_companies_json': json.dumps(groups_to_companies),
#         'companies_departments_json': json.dumps(companies_to_departments),
#     }
#     return render(request, 'add_user.html', context)


# @require_POST
# @login_required
# def update_user(request, user_id):
#     """
#     Handles inline user updates from the user table via AJAX.
#     """
#     if not (request.user.is_superuser or request.user.access_rights == 'admin'):
#         return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

#     user_to_update = get_object_or_404(CustomUser, id=user_id)

#     # Permission check: Admins can only edit users within their own groups
#     if not request.user.is_superuser and request.user.access_rights == 'admin':
#         admin_groups = request.user.groups.all()
#         if not user_to_update.groups.filter(id__in=admin_groups.values_list('id', flat=True)).exists():
#             return JsonResponse({'success': False, 'error': 'Permission denied to edit this user'}, status=403)

#     # Update non-M2M fields
#     user_to_update.username = request.POST.get('username', user_to_update.username)
#     user_to_update.first_name = request.POST.get('first_name', user_to_update.first_name)
#     user_to_update.last_name = request.POST.get('last_name', user_to_update.last_name)
#     user_to_update.email = request.POST.get('email', user_to_update.email)
#     user_to_update.access_rights = request.POST.get('access_rights', user_to_update.access_rights)
#     user_to_update.save()

#     # Update M2M fields
#     group_ids = request.POST.getlist('groups[]')
#     user_to_update.groups.set(group_ids)

#     company_ids = request.POST.getlist('companies[]')
#     user_to_update.companies.set(company_ids)

#     department_ids = request.POST.getlist('departments[]')
#     user_to_update.departments.set(department_ids)
    
#     # Prepare JSON response with updated data
#     response_data = {
#         'success': True,
#         'username': user_to_update.username,
#         'full_name': user_to_update.get_full_name(),
#         'email': user_to_update.email,
#         'groups': ", ".join(g.name for g in user_to_update.groups.all()) or '-',
#         'companies': ", ".join(c.name for c in user_to_update.companies.all()) or '-',
#         'departments': ", ".join(d.name for d in user_to_update.departments.all()) or '-',
#         'access_rights': user_to_update.get_access_rights_display(),
#     }
#     return JsonResponse(response_data)


# @require_POST
# @login_required
# def delete_user(request, user_id):
#     """
#     Deletes a user.
#     """
#     if not (request.user.is_superuser or request.user.access_rights == 'admin'):
#         messages.error(request, "You do not have permission to perform this action.")
#         return redirect('add_user')

#     user_to_delete = get_object_or_404(CustomUser, id=user_id)

#     if user_to_delete == request.user:
#         messages.error(request, "You cannot delete your own account.")
#         return redirect('add_user')
    
#     # Add permission check for admins similar to the update view if needed
#     if not request.user.is_superuser and request.user.access_rights == 'admin':
#         admin_groups = request.user.groups.all()
#         if not user_to_delete.groups.filter(id__in=admin_groups.values_list('id', flat=True)).exists():
#             messages.error(request, "You do not have permission to delete this user.")
#             return redirect('add_user')

#     username = user_to_delete.username
#     user_to_delete.delete()
#     messages.success(request, f"User '{username}' has been deleted.")
#     return redirect('add_user')

# =============================
# PASSWORD MANAGEMENT
# =============================

@login_required
def change_password(request):
    if request.method == 'POST':
        form = PasswordChangeForm(user=request.user, data=request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Important!
            messages.success(request, 'Your password was changed successfully.')
            return redirect('change_password')
        else:
            messages.error(request, 'Please correct the error below.')
    else:
        form = PasswordChangeForm(user=request.user)
    return render(request, 'change_password.html', {'form': form})


# =============================
# HIERARCHY MANAGEMENT
# =============================

# --- USER MANAGEMENT ---

@login_required
def add_user(request):
    user = request.user
    if not (user.is_superuser or user.access_rights == 'admin'):
        messages.error(request, "Permission denied.")
        return redirect('dashboard_view') 

    if request.method == 'POST':
        form = CustomUserForm(request.POST, request_user=request.user)
        if form.is_valid():
            new_user = form.save(commit=False)
            
            # Auto-assign group for admins creating users
            if not user.is_superuser and user.company_group:
                new_user.company_group = user.company_group
                
            new_user.save()
            form.save_m2m() 
            messages.success(request, f"User '{new_user.username}' created.")
            return redirect('add_user')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = CustomUserForm(request_user=user)

    # --- DATA PREP FOR TABLE ---
    if user.is_superuser:
        users = CustomUser.objects.all().order_by('username')
        groups = CompanyGroup.objects.all()
    elif user.access_rights == 'admin' and user.company_group:
        groups = CompanyGroup.objects.filter(id=user.company_group.id)
        users = CustomUser.objects.filter(company_group=user.company_group).distinct().order_by('username')
    else:
        users = CustomUser.objects.none()
        groups = CompanyGroup.objects.none()

    context = {
        'form': form,
        'users': users,
        'groups': groups,
        'all_groups_json': json.dumps(list(groups.values('id', 'name'))),
    }
    return render(request, 'add_user.html', context)


@require_POST
@login_required
def update_user(request, user_id):
    """
    Handles inline user updates via AJAX. 
    Enforces One-Group policy and security checks.
    """
    # 1. Base Permission Check
    if not (request.user.is_superuser or request.user.access_rights == 'admin'):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    user_to_update = get_object_or_404(CustomUser, id=user_id)

    # 2. Group Scope Check (For Admins)
    # An Admin can only edit users within their specific Company Group
    if not request.user.is_superuser:
        if user_to_update.company_group != request.user.company_group:
            return JsonResponse({'success': False, 'error': 'Permission denied: User belongs to a different group.'}, status=403)

    # 3. Update Standard Fields
    # Use .get() to only update if provided, otherwise keep existing
    user_to_update.first_name = request.POST.get('first_name', user_to_update.first_name)
    user_to_update.last_name = request.POST.get('last_name', user_to_update.last_name)
    user_to_update.email = request.POST.get('email', user_to_update.email)
    
    # Validate Access Rights (Prevent Admins from promoting themselves or others to Superuser if you had that role)
    new_access_rights = request.POST.get('access_rights', user_to_update.access_rights)
    if new_access_rights in dict(CustomUser._meta.get_field('access_rights').choices):
        user_to_update.access_rights = new_access_rights

    # 4. Update Group (Superuser Only)
    # The form field in forms.py is named 'company_group', but AJAX might send 'group' or 'company_group'
    new_group_id = request.POST.get('company_group') or request.POST.get('group')
    
    if request.user.is_superuser and new_group_id:
        # Only superusers can move a user to a different group
        user_to_update.company_group_id = new_group_id
    
    # Save basic changes
    user_to_update.save()

    # 5. Prepare Response
    response_data = {
        'success': True,
        'username': user_to_update.username,
        'full_name': user_to_update.get_full_name(),
        'email': user_to_update.email,
        # Display the single group name
        'groups': user_to_update.company_group.name if user_to_update.company_group else '-',
        'access_rights': user_to_update.get_access_rights_display(),
    }
    return JsonResponse(response_data)


@require_POST
@login_required
def delete_user(request, user_id):
    """
    Deletes a user. Enforces strict group boundaries.
    """
    if not (request.user.is_superuser or request.user.access_rights == 'admin'):
        messages.error(request, "Permission denied.")
        return redirect('add_user')

    user_to_delete = get_object_or_404(CustomUser, id=user_id)

    # Prevent self-deletion
    if user_to_delete == request.user:
        messages.error(request, "You cannot delete your own account.")
        return redirect('add_user')
    
    # --- PERMISSION CHECK ---
    # If not superuser, Admin can ONLY delete users in their own company_group
    if not request.user.is_superuser:
        if user_to_delete.company_group != request.user.company_group:
            messages.error(request, "Permission denied: You cannot delete a user from another group.")
            return redirect('add_user')

    username = user_to_delete.username
    user_to_delete.delete()
    messages.success(request, f"User '{username}' has been deleted.")
    return redirect('add_user')







from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Sum
from django.contrib import messages
from django.utils import timezone
from django.http import HttpResponse
import pandas as pd
import io
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

# Make sure to import CreditRequest here
from .models import CompanyGroup, GroupLicense, ProcessingHistory, CreditRequest

# @login_required
# def license_view(request):
#     user = request.user
#     # FIX 1: Use the custom company_group, not standard Django groups
#     user_group = user.company_group 

#     # --- HANDLE CREDIT REQUEST POST ---
#     if request.method == "POST" and 'request_credit_amount' in request.POST:
#         if user_group:
#             try:
#                 amount = int(request.POST.get('request_credit_amount'))
#                 if amount > 0:
#                     CreditRequest.objects.create(
#                         group=user_group,
#                         requested_rows=amount
#                     )
#                     messages.success(request, f"Request for {amount} credits sent to admin.")
#                 else:
#                     messages.error(request, "Please enter a valid amount.")
#             except ValueError:
#                 messages.error(request, "Invalid amount entered.")
#         return redirect('license_view')
#     # ----------------------------------

#     # Defaults for no license/group found
#     context = {
#         'current_expiry': "N/A",
#         'files_remaining': 0, 
#         'logs': [],
#         'license_expired': True,
#         'user_group': user_group # Added for template check
#     }

#     if user_group:
#         # 1. Fetch License Info
#         try:
#             license_obj = GroupLicense.objects.get(group=user_group)
            
#             # Format expiry for display (DD-MM-YYYY HH:MM)
#             context['current_expiry'] = license_obj.expiry_date.strftime("%d-%m-%Y 23:59")
            
#             # Check if expired
#             context['license_expired'] = (license_obj.expiry_date < timezone.now().date()) or (not license_obj.is_active)

#             # 2. Calculate Usage
#             # FIX 2: Changed to .aggregate(Sum) to match the Superuser Manager logic.
#             # This ensures both views show the same "Used" vs "Limit" numbers.
#             total_used = ProcessingHistory.objects.filter(group=user_group).aggregate(total=Sum('rows_processed'))['total'] or 0
            
#             # Calculate remaining
#             remaining = license_obj.row_limit - total_used
#             context['files_remaining'] = max(0, remaining) 

#         except GroupLicense.DoesNotExist:
#             context['current_expiry'] = "No License Found"
        
#         # 3. Fetch History (Logs)
#         history_qs = ProcessingHistory.objects.filter(group=user_group).order_by('-timestamp')
        
#         formatted_logs = []
#         for item in history_qs:
#             # Parse the combined filename "f1 | f2"
#             parts = item.input_filename.split(' | ')
#             f1 = parts[0]
#             f2 = parts[1] if len(parts) > 1 else ""
            
#             formatted_logs.append({
#                 "timestamp": item.timestamp.strftime("%d-%m-%Y %H:%M:%S"),
#                 "file1": f1,
#                 "file2": f2,
#                 "user": item.user.username if item.user else "Unknown" 
#             })
            
#         context['logs'] = formatted_logs

#     return render(request, 'license.html', context)

# @login_required
# def download_history(request):
#     """
#     Exports processing history from Django DB to Excel, matching old Flask format.
#     """
#     user = request.user
#     user_group = user.company_group # Updated to use company_group based on license_view logic
    
#     if not user_group:
#         return redirect('license_view')

#     # 1. Get Logs from DB
#     history_qs = ProcessingHistory.objects.filter(group=user_group).order_by('-timestamp')
    
#     if not history_qs.exists():
#         messages.warning(request, "No processing history found.")
#         return redirect('license_view')

#     try:
#         data = []
#         for item in history_qs:
#             # Split filename back apart
#             parts = item.input_filename.split(' | ')
#             f1 = parts[0]
#             f2 = parts[1] if len(parts) > 1 else ""
            
#             data.append({
#                 'timestamp': item.timestamp.strftime("%d-%m-%Y %H:%M:%S"),
#                 'file1': f1,
#                 'file2': f2
#             })

#         # 2. Convert to DataFrame
#         df = pd.DataFrame(data)
        
#         # 3. Filter and Rename Columns
#         df = df[['timestamp', 'file1', 'file2']]
#         df.columns = ['Timestamp', 'File 1 Name', 'File 2 Name']

#         # 4. Add "Sr No."
#         df.insert(0, 'Sr No.', range(1, len(df) + 1))

#         # 5. Generate Excel
#         output = io.BytesIO()
#         with pd.ExcelWriter(output, engine='openpyxl') as writer:
#             sheet_name = 'Processing History'
#             df.to_excel(writer, index=False, sheet_name=sheet_name)
            
#             worksheet = writer.sheets[sheet_name]

#             # --- HEADER STYLING (Preserved from Flask) ---
#             header_font = Font(name='Calibri', size=11, color="FFFFFF", bold=True)
#             header_fill = PatternFill(start_color="000c66", end_color="000c66", fill_type="solid")
#             header_alignment = Alignment(horizontal='center', vertical='center')

#             for cell in worksheet[1]:
#                 cell.font = header_font
#                 cell.fill = header_fill
#                 cell.alignment = header_alignment

#             # --- AUTO-ADJUST COLUMN WIDTH ---
#             for i, col in enumerate(df.columns, 1):
#                 max_length = 0
#                 column = df.iloc[:, i-1] 
                
#                 # Check header length
#                 val = col
#                 if val: max_length = max(max_length, len(str(val)))
                
#                 # Check data length (sample first 50 rows)
#                 for cell in column.head(50):
#                     if cell: max_length = max(max_length, len(str(cell)))
                
#                 col_letter = get_column_letter(i)
#                 worksheet.column_dimensions[col_letter].width = max_length + 3

#         output.seek(0)
        
#         filename = f"History_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
#         response = HttpResponse(
#             output.getvalue(),
#             content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#         )
#         response['Content-Disposition'] = f'attachment; filename={filename}'
#         return response

#     except Exception as e:
#         print(f"Error generating report: {e}")
#         return redirect('license_view')

# Security check: User must be a superuser (is_superuser=True). 
def is_superuser_check(user):
    return user.is_superuser




from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.contrib import messages
from django.utils import timezone
from .models import GroupLicense, ProcessingHistory, CreditRequest, CustomUser

@login_required
def license_view(request):
    user = request.user
    user_group = user.company_group

    # --- HANDLE CREDIT REQUEST ---
    if request.method == "POST" and 'request_credit_amount' in request.POST:
        if user_group:
            try:
                amount = int(request.POST.get('request_credit_amount'))
                if amount > 0:
                    CreditRequest.objects.create(group=user_group, requested_rows=amount)
                    messages.success(request, f"Request for {amount} credits sent to admin.")
                else:
                    messages.error(request, "Invalid amount.")
            except ValueError:
                messages.error(request, "Invalid input.")
        return redirect('license_view')

    context = {
        'current_expiry': "N/A", 'files_remaining': 0, 'logs': [], 
        'user_stats': [], 'license_expired': True, 'user_group': user_group,
        'all_users': [],
        # Persist filters
        'filter_user': request.GET.get('user', ''),
        'filter_date_from': request.GET.get('date_from', ''),
        'filter_date_to': request.GET.get('date_to', ''),
        'sort_by': request.GET.get('sort', 'timestamp'), # Default sort
        'direction': request.GET.get('direction', 'desc'), # Default direction
    }

    if user_group:
        # 1. License Info
        try:
            license_obj = GroupLicense.objects.get(group=user_group)
            context['current_expiry'] = license_obj.expiry_date.strftime("%d-%m-%Y")
            context['license_expired'] = (license_obj.expiry_date < timezone.now().date()) or (not license_obj.is_active)
            total_used = ProcessingHistory.objects.filter(group=user_group).aggregate(total=Sum('rows_processed'))['total'] or 0
            context['files_remaining'] = max(0, license_obj.row_limit - total_used)
        except GroupLicense.DoesNotExist:
            context['current_expiry'] = "No License"

        # 2. User Stats
        context['user_stats'] = ProcessingHistory.objects.filter(group=user_group).values(
            'user__username', 'user__first_name', 'user__last_name'
        ).annotate(total_usage=Sum('rows_processed')).order_by('-total_usage')

        # 3. History with Filters & Sort
        history_qs = ProcessingHistory.objects.filter(group=user_group)

        # Filters
        if context['filter_user']:
            history_qs = history_qs.filter(user__username=context['filter_user'])
        if context['filter_date_from']:
            history_qs = history_qs.filter(timestamp__date__gte=context['filter_date_from'])
        if context['filter_date_to']:
            history_qs = history_qs.filter(timestamp__date__lte=context['filter_date_to'])

        # Sorting Logic
        sort_col = context['sort_by']
        direction = context['direction']
        
        # Map URL params to DB fields
        ordering_map = {
            'timestamp': 'timestamp',
            'user': 'user__username',
            'file1': 'input_filename', # Assuming file1 is start of string
            'rows': 'rows_processed'
        }
        
        db_sort = ordering_map.get(sort_col, 'timestamp')
        if direction == 'desc':
            db_sort = f'-{db_sort}'
            
        history_qs = history_qs.order_by(db_sort)

        # Formatting for View
        formatted_logs = []
        for item in history_qs:
            parts = item.input_filename.split(' | ')
            f1 = parts[0]
            f2 = parts[1] if len(parts) > 1 else "-"
            
            user_display = "System"
            if item.user:
                user_display = item.user.get_full_name() or item.user.username

            formatted_logs.append({
                "timestamp": item.timestamp, # Pass object for flexible formatting
                "group_name": user_group.name,
                "user": user_display,
                "file1": f1,
                "file2": f2,
                "rows": item.rows_processed
            })
            
        context['logs'] = formatted_logs
        context['all_users'] = CustomUser.objects.filter(company_group=user_group)

    return render(request, 'license.html', context)





# Add these imports if they are missing
from django.db.models import Sum
from django.contrib.auth.decorators import user_passes_test
import pandas as pd
import io
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ==========================================
#  NEW: LICENSE MANAGER REPORT VIEW
# ==========================================
@user_passes_test(lambda u: u.is_superuser)
def download_license_report(request):
    """
    Generates a system-wide report for the Superuser.
    Columns: Group, Admin, Limit, Used, Balance, Expiry (dd-mm-yyyy), Status.
    """
    # 1. Fetch all groups with optimization
    groups = CompanyGroup.objects.all().select_related('license', 'created_by')

    data = []
    for group in groups:
        # Calculate Total Used Rows directly from History
        total_used = ProcessingHistory.objects.filter(group=group).aggregate(total=Sum('rows_processed'))['total'] or 0
        
        # Default values if no license exists
        limit = 0
        expiry_str = "N/A"
        status = "No License"
        
        if hasattr(group, 'license'):
            lic = group.license
            limit = lic.row_limit
            status = "Active" if lic.is_active else "Inactive"
            
            # --- DATE FORMAT LOGIC (dd-mm-yyyy) ---
            if lic.expiry_date:
                expiry_str = lic.expiry_date.strftime('%d-%m-%Y')
        
        data.append({
            'Group Name': group.name,
            'Group Admin': group.created_by.email if group.created_by else "N/A",
            'Total Limit': limit,
            'Credits Used': total_used,
            'Balance': limit - total_used,
            'Expiry Date': expiry_str,  # Formatted as dd-mm-yyyy
            'Status': status
        })

    # 2. Create DataFrame
    df = pd.DataFrame(data)
    if not df.empty:
        # Add Serial Number
        df.insert(0, 'Sr No.', range(1, len(df) + 1))
    
    # 3. Generate Excel
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        sheet_name = 'License Report'
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        
        # --- STYLING ---
        worksheet = writer.sheets[sheet_name]
        
        # Header Style: Blue background, White text, Centered
        header_font = Font(name='Calibri', size=11, color="FFFFFF", bold=True)
        header_fill = PatternFill(start_color="000c66", end_color="000c66", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Auto-width columns
        for col in worksheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try: max_len = max(max_len, len(str(cell.value)))
                except: pass
            worksheet.column_dimensions[col_letter].width = max_len + 3

    output.seek(0)
    filename = f"License_Master_Report_{datetime.now().strftime('%d-%m-%Y')}.xlsx"
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response


from django.http import JsonResponse
from django.utils import timezone
from django.core.exceptions import ObjectDoesNotExist
from .models import GroupLicense, ProcessingHistory  # Ensure models are imported
import math


import math
from django.db.models import Sum

def check_license_validity(user, original_file1, original_file2, row_count_f1, row_count_f2):
    user_group = getattr(user, 'company_group', None)
    
    if not user_group:
        # ADDED , 0
        return False, JsonResponse({'status': 'error', 'message': 'User does not belong to any Company Group.'}, status=403), 0

    unique_key_str = f"{original_file1}|{original_file2}_{row_count_f1}_{row_count_f2}"
    print(f"Checking License for Key: {unique_key_str}")

    try:
        license_obj = GroupLicense.objects.get(group=user_group)

        is_row_based = getattr(license_obj, 'is_row_based', False)
        is_file_based = getattr(license_obj, 'is_file_based', True)

        max_rows = max(row_count_f1, row_count_f2)
        
        if is_row_based:
            ratio = license_obj.rows_per_credit or 500 
            credits_required = math.ceil(max_rows / float(ratio))
            if credits_required == 0:
                credits_required = 1
        elif is_file_based:
            credits_required = 1
        else:
            credits_required = 1

        if license_obj.expiry_date < timezone.now().date():
            # ADDED , credits_required
            return False, JsonResponse({'status': 'error', 'message': 'License Expired.'}, status=403), credits_required
        
        existing_history_record = ProcessingHistory.objects.filter(
            group=user_group, 
            unique_key=unique_key_str
        ).first()

        if existing_history_record:
            print("Record exists - Skipping credit check")
            # ADDED , 0
            return True, existing_history_record, 0
        else:
            credits_used = ProcessingHistory.objects.filter(group=user_group).aggregate(total=Sum('rows_processed'))['total'] or 0
            
            if credits_used + credits_required > license_obj.row_limit:
                # ADDED , credits_required
                return False, JsonResponse({
                    'status': 'error', 
                    'message': f'Credit Limit Reached. Required: {credits_required}, Remaining: {license_obj.row_limit - credits_used}.'
                }, status=403), credits_required
            
            # ADDED , credits_required
            return True, None, credits_required

    except ObjectDoesNotExist:
        # ADDED , 0
        return False, JsonResponse({'status': 'error', 'message': 'No valid license found.'}, status=403), 0




from itertools import groupby
from operator import itemgetter

@login_required
def download_history(request):
    user = request.user

    # --- 1. Filter Logic ---
    if user.is_superuser:
        history_qs = ProcessingHistory.objects.all().select_related('group', 'user').order_by('group__name', '-timestamp')
    elif user.company_group:
        history_qs = ProcessingHistory.objects.filter(group=user.company_group).select_related('group', 'user').order_by('group__name', '-timestamp')
    else:
        return redirect('license_view')

    if not history_qs.exists():
        messages.warning(request, "No processing history found.")
        return redirect('license_view')

    try:
        # ==========================================
        # SHEET 1: DETAILED HISTORY (Flat Data)
        # ==========================================
        data_sheet1 = []
        for item in history_qs:
            parts = item.input_filename.split(' | ')
            f1 = parts[0]
            f2 = parts[1] if len(parts) > 1 else ""

            user_display = "System/Unknown"
            if item.user:
                user_display = item.user.get_full_name() or item.user.username

            data_sheet1.append({
                'timestamp': item.timestamp.strftime("%d-%m-%Y %H:%M:%S"),
                'group': item.group.name if item.group else "N/A",
                'user': user_display,
                'file1': f1,
                'file2': f2,
                'rows': item.rows_processed
            })

        df_details = pd.DataFrame(data_sheet1)
        df_details = df_details[['timestamp', 'group', 'user', 'file1', 'file2', 'rows']]
        df_details.columns = ['Timestamp', 'Group Name', 'User', 'File 1 Name', 'File 2 Name', 'Credits Used']
        df_details.insert(0, 'Sr No.', range(1, len(df_details) + 1))

        # ==========================================
        # SHEET 2: TALLY FORMAT SUMMARY (Hierarchical)
        # ==========================================
        
        # 1. Fetch Aggregated Data
        summary_qs = history_qs.values(
            'group__name', 
            'user__username', 
            'user__first_name', 
            'user__last_name'
        ).annotate(
            total_usage=Sum('rows_processed')
        ).order_by('group__name', '-total_usage')

        # 2. Convert to List for Grouping
        summary_list = list(summary_qs)
        
        tally_data = []
        grand_total = 0
        
        # 3. Group by 'group__name'
        # sort_key is required for groupby
        summary_list.sort(key=itemgetter('group__name'))
        
        for group_name, items in groupby(summary_list, key=itemgetter('group__name')):
            # --- A. Group Header ---
            tally_data.append({'Particulars': f"Group: {group_name}", 'Credits Used': None, 'IsHeader': True})
            
            group_total = 0
            
            # --- B. User Rows ---
            for row in items:
                # Construct Name
                f = row['user__first_name'] or ""
                l = row['user__last_name'] or ""
                full_name = f"{f} {l}".strip() or row['user__username'] or "System"
                
                usage = row['total_usage'] or 0
                group_total += usage
                
                # Add indentation (4 spaces) for Tally look
                tally_data.append({'Particulars': f"    {full_name}", 'Credits Used': usage, 'IsHeader': False})
            
            # --- C. Group Subtotal ---
            tally_data.append({'Particulars': f"Total {group_name}", 'Credits Used': group_total, 'IsTotal': True})
            
            # --- D. Spacer Row ---
            tally_data.append({'Particulars': "", 'Credits Used': None, 'IsHeader': False})
            
            grand_total += group_total

        # --- E. Grand Total ---
        tally_data.append({'Particulars': "GRAND TOTAL", 'Credits Used': grand_total, 'IsGrandTotal': True})

        # Create DataFrame (drop the helper flags later)
        df_tally = pd.DataFrame(tally_data)

        # ==========================================
        # WRITE TO EXCEL
        # ==========================================
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            
            # --- Write Sheet 1 ---
            df_details.to_excel(writer, index=False, sheet_name='Processing History')
            
            # --- Write Sheet 2 (Drop helper columns first) ---
            # We filter columns but keep the dataframe with flags for styling logic below
            export_tally = df_tally[['Particulars', 'Credits Used']]
            export_tally.to_excel(writer, index=False, sheet_name='User Usage Summary')

            # ==========================================
            # STYLING
            # ==========================================
            workbook = writer.book
            
            # --- STYLE SHEET 1 (Standard Table) ---
            ws1 = workbook['Processing History']
            header_fill = PatternFill(start_color="000c66", end_color="000c66", fill_type="solid")
            header_font = Font(name='Calibri', size=11, color="FFFFFF", bold=True)
            
            for cell in ws1[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                
            for col in ws1.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try: max_len = max(max_len, len(str(cell.value)))
                    except: pass
                ws1.column_dimensions[col_letter].width = max_len + 3

            # --- STYLE SHEET 2 (Tally Format) ---
            ws2 = workbook['User Usage Summary']
            
            # Style the Main Header Row
            for cell in ws2[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            ws2.column_dimensions['A'].width = 50  # Particulars
            ws2.column_dimensions['B'].width = 20  # Credits

            # Loop through rows to apply Conditional Formatting based on our data structure
            # Data starts at row 2 (index 1 in 0-based list)
            
            bold_font = Font(name='Calibri', bold=True)
            red_bold_font = Font(name='Calibri', bold=True, color="990000")
            top_border = Border(top=Side(style='thin'))
            double_bottom = Border(bottom=Side(style='double'), top=Side(style='thin'))

            for i, row_data in enumerate(tally_data):
                excel_row_idx = i + 2 # Header is 1, so data starts at 2
                
                # Check flags from our dictionary
                is_header = row_data.get('IsHeader', False)
                is_total = row_data.get('IsTotal', False)
                is_grand = row_data.get('IsGrandTotal', False)
                
                if is_header:
                    # Group Name: Bold
                    ws2.cell(row=excel_row_idx, column=1).font = bold_font
                    
                elif is_total:
                    # Subtotal: Bold + Top Border + Right Align text
                    cell_name = ws2.cell(row=excel_row_idx, column=1)
                    cell_val = ws2.cell(row=excel_row_idx, column=2)
                    
                    cell_name.font = bold_font
                    cell_name.alignment = Alignment(horizontal='right') 
                    cell_name.border = top_border
                    
                    cell_val.font = bold_font
                    cell_val.border = top_border

                elif is_grand:
                    # Grand Total: Red Bold + Double Bottom Border
                    cell_name = ws2.cell(row=excel_row_idx, column=1)
                    cell_val = ws2.cell(row=excel_row_idx, column=2)
                    
                    cell_name.font = red_bold_font
                    cell_name.alignment = Alignment(horizontal='right')
                    cell_name.border = double_bottom
                    
                    cell_val.font = red_bold_font
                    cell_val.border = double_bottom

        output.seek(0)
        filename = f"History_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'
        return response

    except Exception as e:
        print(f"Error generating report: {e}")
        return redirect('license_view')




# Ensure you import your models: CreditRequest, GroupLicense, CompanyGroup

@user_passes_test(lambda u: u.is_superuser)
def license_manager_view(request):
    """
    View for Superusers to manage Group Licenses and Approve/Reject Requests.
    """
    if request.method == "POST":
        # === HANDLE CREDIT REQUEST APPROVAL/REJECTION ===
        if 'action_type' in request.POST and request.POST['action_type'] == 'credit_request':
            request_id = request.POST.get('request_id')
            action = request.POST.get('request_action') # 'approve' or 'reject'
            
            try:
                credit_req = CreditRequest.objects.get(pk=request_id)
                
                if action == 'approve':
                    # Update License
                    group = credit_req.group
                    license_obj, created = GroupLicense.objects.get_or_create(
                        group=group,
                        defaults={'row_limit': 0, 'expiry_date': timezone.now().date()}
                    )
                    
                    # --- NEW LOGIC START: Update Expiry to 1 Year from Today ---
                    today = timezone.now().date()
                    try:
                        # Try to replace the year with next year (works for most dates)
                        new_expiry_date = today.replace(year=today.year + 1)
                    except ValueError:
                        # Handle Leap Year edge case (Feb 29 becomes Feb 28 next year)
                        new_expiry_date = today.replace(year=today.year + 1, day=28)
                    
                    license_obj.expiry_date = new_expiry_date
                    # --- NEW LOGIC END ---

                    license_obj.row_limit += credit_req.requested_rows
                    license_obj.save()
                    
                    credit_req.status = 'APPROVED'
                    credit_req.processed_at = timezone.now()
                    credit_req.save()
                    messages.success(request, f"Approved {credit_req.requested_rows} credits for {group.name}. Expiry extended to {new_expiry_date}.")
                    
                elif action == 'reject':
                    credit_req.status = 'REJECTED'
                    credit_req.processed_at = timezone.now()
                    credit_req.save()
                    messages.warning(request, f"Rejected credit request for {credit_req.group.name}.")
                    
            except CreditRequest.DoesNotExist:
                messages.error(request, "Request not found.")
            except Exception as e:
                messages.error(request, f"Error processing request: {e}")
                
            return redirect('license_manager')

        # === HANDLE MANUAL LICENSE UPDATE ===
        else:
            group_id = request.POST.get('group_id')
            new_row_limit = request.POST.get('row_limit')
            new_expiry_date = request.POST.get('expiry_date')
            is_active = request.POST.get('is_active') == 'on'
            
            # --- 3. GET THE LICENSE TYPE ---
            license_type = request.POST.get('license_type')
            rows_per_credit = request.POST.get('rows_per_credit', 500)
            # -------------------------------

            try:
                group = CompanyGroup.objects.get(pk=group_id)
                
                # Get or Create the license if it doesn't exist
                license_obj, created = GroupLicense.objects.get_or_create(
                    group=group,
                    defaults={
                        'expiry_date': timezone.now().date(),
                        'row_limit': 0
                    }
                )

                # Update fields
                if new_row_limit:
                    license_obj.row_limit = int(new_row_limit)
                
                if new_expiry_date:
                    license_obj.expiry_date = new_expiry_date
                
                license_obj.is_active = is_active

                # --- 4. SAVE THE LICENSE TYPE ---
                if license_type == 'row':
                    license_obj.is_row_based = True
                    license_obj.is_file_based = False
                    # Store the ratio provided by the admin
                    license_obj.rows_per_credit = int(request.POST.get('rows_per_credit', 500))
                else:
                    license_obj.is_row_based = False
                    license_obj.is_file_based = True
                    # NEW: Set to None (NULL) because it isn't needed for File Based
                    license_obj.rows_per_credit = None
                # --------------------------------

                license_obj.save()

                messages.success(request, f"License for '{group.name}' updated successfully.")

            except CompanyGroup.DoesNotExist:
                messages.error(request, "Company Group not found.")
            except ValueError:
                messages.error(request, "Invalid input data.")
            except Exception as e:
                messages.error(request, f"An error occurred: {str(e)}")
            
            return redirect('license_manager')

    # GET Request: Prepare dashboard data
    groups = CompanyGroup.objects.all().select_related('license').prefetch_related('processing_history')
    
    # Fetch Pending Requests
    pending_requests = CreditRequest.objects.filter(status='PENDING').select_related('group').order_by('-created_at')

    dashboard_data = []

    for group in groups:
        # Calculate Total Used Rows from History
        total_used = group.processing_history.aggregate(total=Sum('rows_processed'))['total'] or 0
        
        # Get License Data safely
        if hasattr(group, 'license'):
            limit = group.license.row_limit
            expiry = group.license.expiry_date
            active = group.license.is_active
            # Calculate Balance
            balance = limit - total_used
        else:
            limit = 0
            expiry = None
            active = False
            balance = 0

        dashboard_data.append({
            'group_obj': group,
            'total_used': total_used,
            'limit': limit,
            'balance': balance,
            'expiry': expiry,
            'active': active,
            'has_license': hasattr(group, 'license')
        })

    context = {
        'dashboard_data': dashboard_data,
        'pending_requests': pending_requests,
        'page_title': 'Global License Manager'
    }

    return render(request, 'license_manager.html', context)




















# --- HELPER FUNCTIONS (Must be present) ---

DB_PATH = ':memory:'

def convert_to_date_only(df):
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = df[col].dt.strftime('%d-%m-%Y')
    return df

def clean_string(input_str, words_to_remove=None):
    if not isinstance(input_str, str): return ""
    if words_to_remove is None:
        words_to_remove = {"RTGS", "EMI", "Cr", "FT", "CHQ", "Paid", "TRANSFER", "NEFT", "PVT","LTD", "CO", "CREDIT","CARD","HUF", "WITH", "SELF", "Bank", "Reference", "DEP", "NO.", "NO", "-", "/", "//", "@", ".", "_", ":", "(",")","IN"}
    pattern = r'\b(?:' + '|'.join(map(re.escape, words_to_remove)) + r')\b'
    cleaned_str = re.sub(pattern, '', input_str, flags=re.IGNORECASE)  
    cleaned_str = re.sub(r'\s+', ' ', cleaned_str).strip().lower()
    return cleaned_str

def common_word_match(str1, str2, threshold):
    similarity_score = fuzz.token_set_ratio(str(str1), str(str2))
    return similarity_score >= threshold

def get_table_column_types(conn, table_name):
    query = f"PRAGMA table_info('{table_name}')"
    cursor = conn.execute(query)
    return {row[1]: row[2] for row in cursor.fetchall()}

def get_table_columns(conn, table_name):
    query = f"PRAGMA table_info('{table_name}')"
    cursor = conn.execute(query)
    return [row[1] for row in cursor.fetchall()]

def filter_dataframe_columns(df, selected_cols):
    if not selected_cols: return df
    existing_cols = [col for col in selected_cols if col in df.columns]
    return df[existing_cols] if existing_cols else df

def generate_difference_query(f1_raw_cols, f2_raw_cols, f1_type_map, f2_type_map, 
                              convert_type, diff_value, diff_range):
    """
    Generates SQL for Difference Reconciliation.
    First N-1 columns are Exact Match (respecting Number Format).
    Last Column is the Numeric Match with Tolerance.
    Adds a 'Difference' column to the output.
    """
    f1_table = 'file1_table'
    f2_table = 'file2_table'
    join_conditions = []

    # 1. Handle Exact Match Columns (All except the last one)
    num_exact_cols = len(f1_raw_cols) - 1
    
    for i in range(num_exact_cols):
        raw_col_1 = f1_raw_cols[i]
        raw_col_2 = f2_raw_cols[i]
        
        f1_col_sql = f'{f1_table}."{raw_col_1}"'
        f2_col_sql = f'{f2_table}."{raw_col_2}"'
        
        f1_type = f1_type_map.get(raw_col_1, 'TEXT')
        f2_type = f2_type_map.get(raw_col_2, 'TEXT')
        
        if f1_type == f2_type and f1_type in ('INTEGER', 'REAL', 'FLOAT', 'DOUBLE', 'DECIMAL', 'BIGINT', 'NUMERIC', 'HUGEINT', 'SMALLINT', 'TINYINT', 'UBIGINT', 'UINTEGER', 'USMALLINT', 'UTINYINT'):
            # --- APPLIED NUMBER FORMAT LOGIC TO IDENTIFIERS TOO ---
            if convert_type == 'roundoff':
                join_conditions.append(f"ROUND({f1_col_sql}, 0) = ROUND({f2_col_sql}, 0)")
            else:
                # Default to 2 decimal places for "Decimal" match to avoid float precision errors
                join_conditions.append(f"ROUND({f1_col_sql}, 2) = ROUND({f2_col_sql}, 2)")
        else:
            join_conditions.append(f"TRIM(LOWER(CAST({f1_col_sql} AS TEXT))) = TRIM(LOWER(CAST({f2_col_sql} AS TEXT)))")

    # 2. Handle The Value Column (The Last One)
    val_col_1 = f1_raw_cols[-1]
    val_col_2 = f2_raw_cols[-1]
    
    v1_sql = f'CAST({f1_table}."{val_col_1}" AS DOUBLE)'
    v2_sql = f'CAST({f2_table}."{val_col_2}" AS DOUBLE)'

    # Apply Number Format Logic for Matching
    if convert_type == 'roundoff':
        v1_final = f"ROUND({v1_sql}, 0)"
        v2_final = f"ROUND({v2_sql}, 0)"
    else: 
        v1_final = f"ROUND({v1_sql}, 2)"
        v2_final = f"ROUND({v2_sql}, 2)"

    # Difference Calculation (File 1 - File 2)
    diff_calc = f"({v1_final} - {v2_final})"
    
    try:
        tolerance = float(diff_value)
    except:
        tolerance = 0.0

    if diff_range == 'positive':
        join_conditions.append(f"{diff_calc} >= 0 AND {diff_calc} <= {tolerance}")
    elif diff_range == 'negative':
        join_conditions.append(f"{diff_calc} <= 0 AND {diff_calc} >= -{tolerance}")
    else: # 'both'
        join_conditions.append(f"ABS({diff_calc}) <= {tolerance}")

    # 3. Construct Final Queries
    join_clause = ' AND '.join(join_conditions)
    
    common_query = f"""
        SELECT {f1_table}.*, {f2_table}.*, {diff_calc} AS "Difference" 
        FROM {f1_table} 
        INNER JOIN {f2_table} ON {join_clause}
    """
    
    f1_not_in = f"""
        SELECT {f1_table}.* FROM {f1_table} 
        LEFT JOIN {f2_table} ON {join_clause} 
        WHERE {f2_table}."__row_index__" IS NULL
    """
    
    f2_not_in = f"""
        SELECT {f2_table}.* FROM {f2_table} 
        LEFT JOIN {f1_table} ON {join_clause} 
        WHERE {f1_table}."__row_index__" IS NULL
    """

    return common_query, f1_not_in, f2_not_in, join_clause






def clean_dataframe_quotes(df):
    """
    Removes leading/trailing single quotes (') and double quotes (") 
    from all text columns in the DataFrame.
    """
    for col in df.select_dtypes(include=['object']).columns:
        # Only apply to string values, leave NaNs/Numbers alone
        df[col] = df[col].apply(lambda x: x.strip("'").strip('"') if isinstance(x, str) else x)
    return df

def process_reco(request):
    return render(request, 'reco.html')





# --- ADD TO HELPER FUNCTIONS ---

# --- HELPER FUNCTIONS ---

def analyze_pandas_mismatch(source_df, target_df, exact_cols, fuzzy_col, target_exact_cols, target_fuzzy_col, threshold):
    """
    Analyzes mismatch and returns detailed info.
    Returns: { row_index: {'col': source_col, 'target_col': target_col, 'seq': sequence_int} }
    """
    fail_map = {}
    
    # Pre-calculate cleaned fuzzy strings for target
    if '__clean_fuzzy__' not in target_df.columns:
        target_df = target_df.copy()
        target_df['__clean_fuzzy__'] = target_df[target_fuzzy_col].apply(clean_string)

    for idx, src_row in source_df.iterrows():
        row_id = src_row.get('__row_index__')
        current_candidates = target_df
        mismatch_found = False
        
        # 1. Check Exact Columns in order
        for i, col in enumerate(exact_cols):
            tgt_col = target_exact_cols[i]
            
            # Filter candidates
            next_candidates = current_candidates[current_candidates[tgt_col] == src_row[col]]
            
            if next_candidates.empty:
                # Failure found at this step
                fail_map[row_id] = {
                    'col': col,
                    'target_col': tgt_col,
                    'seq': i + 1
                }
                mismatch_found = True
                break
            else:
                current_candidates = next_candidates
        
        if mismatch_found: continue
            
        # 2. Check Fuzzy Threshold (Final Step)
        src_clean = src_row.get('__clean_fuzzy__')
        if not isinstance(src_clean, str): src_clean = clean_string(src_row[fuzzy_col])
            
        found_fuzzy_match = False
        for _, tgt_row in current_candidates.iterrows():
            if common_word_match(src_clean, tgt_row['__clean_fuzzy__'], threshold):
                found_fuzzy_match = True
                break
        
        if not found_fuzzy_match:
            fail_map[row_id] = {
                'col': fuzzy_col,
                'target_col': target_fuzzy_col,
                'seq': len(exact_cols) + 1
            }

    return fail_map

def analyze_sql_mismatch(conn, source_table, target_table, source_cols, target_cols, source_unrec_ids):
    """
    SQL Version: Returns details on which column caused the mismatch.
    Returns: { row_index: {'col': source_col, 'target_col': target_col, 'seq': sequence_int} }
    """
    if not source_cols: return {}
    fail_map = {}
    num_cols = len(source_cols)
    
    # Iterate backwards from Last Column down to First
    for i in range(num_cols - 1, -1, -1):
        failing_col_name = source_cols[i]
        target_col_name = target_cols[i]
        sequence_no = i + 1
        
        # If we are at the first column (i=0), anyone not mapped yet failed here.
        if i == 0:
            current_stage_ids = source_unrec_ids - set(fail_map.keys())
            for uid in current_stage_ids:
                fail_map[uid] = {
                    'col': failing_col_name,
                    'target_col': target_col_name,
                    'seq': sequence_no
                }
            break

        # Build Join for columns 0 to i-1 (The ones BEFORE the current check)
        subset_src = source_cols[:i]
        subset_tgt = target_cols[:i]
        conditions = []
        for k in range(len(subset_src)):
            # Simple cast for analysis to catch mismatches
            f1 = f'{source_table}."{subset_src[k]}"'
            f2 = f'{target_table}."{subset_tgt[k]}"'
            conditions.append(f"CAST({f1} AS TEXT) = CAST({f2} AS TEXT)")
        
        join_clause = " AND ".join(conditions)
        
        # Find rows that matched PREVIOUS cols but failed CURRENT col
        # (We query for rows that match previous cols, intersect with our Unreconciled list)
        query = f"""
            SELECT {source_table}."__row_index__" FROM {source_table}
            INNER JOIN {target_table} ON {join_clause}
            WHERE {source_table}."__row_index__" IN ({','.join(map(str, source_unrec_ids))})
        """
        try:
            found_ids = set(conn.execute(query).df()['__row_index__'].tolist())
            for uid in found_ids:
                if uid not in fail_map:
                    fail_map[uid] = {
                        'col': failing_col_name,
                        'target_col': target_col_name,
                        'seq': sequence_no
                    }
        except: pass

    return fail_map

def save_excel_with_highlights(writer, df, sheet_name, highlight_map, output_cols=None):
    """
    Adds 'Sequence No' and 'Matched To' columns, then writes to Excel with Red Highlights.
    highlight_map: { row_id: {'col':..., 'target_col':..., 'seq':...} }
    """
    # 1. Prepare Display Data
    df_display = df.copy()
    
    # 2. Inject New Columns using the Map
    # We create lists to populate the new columns efficiently
    seq_list = []
    matched_to_list = []
    
    # Map row_index to the dataframe index to populate lists
    if '__row_index__' in df_display.columns:
        for val in df_display['__row_index__']:
            if val in highlight_map:
                details = highlight_map[val]
                seq_list.append(details['seq'])
                matched_to_list.append(details['target_col'])
            else:
                seq_list.append(None)
                matched_to_list.append(None)
    else:
        seq_list = [None] * len(df_display)
        matched_to_list = [None] * len(df_display)

    # Add columns to DataFrame
    df_display['Sequence No'] = seq_list
    df_display['Matched To'] = matched_to_list

    # 3. Clean up internal columns
    cols_to_drop = ['__row_index__', '__clean_fuzzy__', '__orig_idx__']
    df_display = df_display.drop(columns=[c for c in cols_to_drop if c in df_display.columns], errors='ignore')
    
    # 4. Handle Column Filtering (Ensure new columns are kept)
    if output_cols:
        # We must explicitly add our new columns to the requested list temporarily
        temp_cols = list(output_cols) + ['Sequence No', 'Matched To']
        # Only keep columns that actually exist
        valid_cols = [c for c in temp_cols if c in df_display.columns]
        df_display = df_display[valid_cols]
    
    # 5. Write to Excel
    df_display.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # 6. Apply Red Highlights
    workbook = writer.book
    sheet = workbook[sheet_name]
    
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(color="9C0006") 

    col_name_to_idx = {name: idx + 1 for idx, name in enumerate(df_display.columns)}
    
    if '__row_index__' not in df.columns: return

    # Loop through original DF to get IDs, but write to Sheet based on new layout
    for i, row in enumerate(df.itertuples(index=False)):
        try:
            internal_id = df.iloc[i]['__row_index__']
        except: continue
            
        if internal_id in highlight_map:
            failing_col = highlight_map[internal_id]['col']
            
            if failing_col in col_name_to_idx:
                col_idx = col_name_to_idx[failing_col]
                excel_row = i + 2 
                
                cell = sheet.cell(row=excel_row, column=col_idx)
                cell.fill = red_fill
                cell.font = red_font


# --- VIEWS ---

@csrf_exempt
@login_required
def upload_files(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'fail', 'message': 'Method not allowed'}, status=405)

    temp_dir = tempfile.gettempdir()
    column_names = {}

    # --- INPUT VARIABLES ---
    # Check if we are continuing from a sheet selection (using existing paths)
    existing_p1 = request.POST.get('existing_file1_path')
    existing_p2 = request.POST.get('existing_file2_path')
    
    # Check if specific sheets were selected
    selected_sheet1 = request.POST.get('sheet_name_file1')
    selected_sheet2 = request.POST.get('sheet_name_file2')

    file1_path = None
    file2_path = None
    original_file1_name = "File1"
    original_file2_name = "File2"

    try:
        # --- STEP 1: HANDLE FILE 1 ---
        if existing_p1 and os.path.exists(existing_p1):
            # Case A: User selected a sheet, re-using uploaded file
            file1_path = existing_p1
            original_file1_name = os.path.basename(existing_p1) # Simplification
        else:
            # Case B: New File Upload
            file1 = request.FILES.get('file1')
            if not file1: return JsonResponse({'status': 'fail', 'message': 'File 1 is missing.'}, status=400)
            
            unique_id = datetime.now().strftime("%Y%m%dH%M%S%f")
            original_file1_name = file1.name
            file1_filename = f"file1_{unique_id}_{secure_filename(file1.name or 'f1.xlsx')}"
            file1_path = os.path.join(temp_dir, file1_filename)

            with open(file1_path, 'wb+') as destination:
                for chunk in file1.chunks():
                    destination.write(chunk)

        # --- STEP 2: HANDLE FILE 2 ---
        if existing_p2 and os.path.exists(existing_p2):
            file2_path = existing_p2
            original_file2_name = os.path.basename(existing_p2)
        else:
            file2 = request.FILES.get('file2')
            if not file2: return JsonResponse({'status': 'fail', 'message': 'File 2 is missing.'}, status=400)

            unique_id = datetime.now().strftime("%Y%m%dH%M%S%f")
            original_file2_name = file2.name
            file2_filename = f"file2_{unique_id}_{secure_filename(file2.name or 'f2.xlsx')}"
            file2_path = os.path.join(temp_dir, file2_filename)

            with open(file2_path, 'wb+') as destination:
                for chunk in file2.chunks():
                    destination.write(chunk)

        # --- STEP 3: CHECK FOR MULTI-SHEET EXCEL (Before Loading Data) ---
        # We only check sheets if this is a NEW upload (no specific sheet selected yet)
        sheets_info = {}
        
        # Check File 1 Sheets
        if not selected_sheet1 and not file1_path.lower().endswith('.csv'):
            xl1 = pd.ExcelFile(file1_path)
            if len(xl1.sheet_names) > 1:
                sheets_info['file1'] = xl1.sheet_names

        # Check File 2 Sheets
        if not selected_sheet2 and not file2_path.lower().endswith('.csv'):
            xl2 = pd.ExcelFile(file2_path)
            if len(xl2.sheet_names) > 1:
                sheets_info['file2'] = xl2.sheet_names

        # IF MULTIPLE SHEETS FOUND, RETURN TO FRONTEND FOR SELECTION
        if sheets_info:
            return JsonResponse({
                'status': 'sheet_selection',
                'message': 'Multiple sheets detected. Please select one.',
                'sheets_data': sheets_info,
                'file1_path': file1_path,
                'file2_path': file2_path,
                'original_file1': original_file1_name,
                'original_file2': original_file2_name
            })


        # --- STEP 4: LOAD DATAFRAMES (Specific Sheet or Default) ---
        
        # Load File 1
        if file1_path.lower().endswith('.csv'):
            df_file1 = pd.read_csv(file1_path)
        else:
            # Load specific sheet if provided, else default (0)
            sheet_to_load = selected_sheet1 if selected_sheet1 else 0
            df_file1 = pd.read_excel(file1_path, sheet_name=sheet_to_load)
        
        # Load File 2
        if file2_path.lower().endswith('.csv'):
            df_file2 = pd.read_csv(file2_path)
        else:
            sheet_to_load = selected_sheet2 if selected_sheet2 else 0
            df_file2 = pd.read_excel(file2_path, sheet_name=sheet_to_load)

        # --- RENAMING LOGIC (CASE INSENSITIVE FIX) ---
        cols1_lower = {c.lower() for c in df_file1.columns}
        cols2_lower = {c.lower() for c in df_file2.columns}
        common_cols_lower = cols1_lower.intersection(cols2_lower)

        if common_cols_lower:
            rename_map_1 = {}
            rename_map_2 = {}
            
            for col in df_file1.columns:
                if col.lower() in common_cols_lower:
                    # CHANGED HERE: Put 'file1_' at the start
                    rename_map_1[col] = f"file1_{col}"
            
            for col in df_file2.columns:
                if col.lower() in common_cols_lower:
                    # CHANGED HERE: Put 'file2_' at the start
                    rename_map_2[col] = f"file2_{col}"
            
            df_file1.rename(columns=rename_map_1, inplace=True)
            df_file2.rename(columns=rename_map_2, inplace=True)
            
            # Save the processed/renamed version back to disk to ensure consistency for the next steps
            if file1_path.lower().endswith('.csv'):
                df_file1.to_csv(file1_path, index=False)
            else:
                df_file1.to_excel(file1_path, index=False)
                
            if file2_path.lower().endswith('.csv'):
                df_file2.to_csv(file2_path, index=False)
            else:
                df_file2.to_excel(file2_path, index=False)

        column_names['File 1'] = df_file1.columns.tolist()
        column_names['File 2'] = df_file2.columns.tolist()

    except Exception as e:
        # cleanup if error and it was a new upload
        if not existing_p1 and os.path.exists(file1_path): os.remove(file1_path)
        if not existing_p2 and os.path.exists(file2_path): os.remove(file2_path)
        import traceback
        traceback.print_exc()
        return JsonResponse({'status': 'fail', 'message': f'Error processing files: {str(e)}'}, status=500)
    
    return JsonResponse({
        'status': 'success',
        'message': 'Files loaded successfully.',
        'columns': column_names,
        'file1_path': file1_path,
        'file2_path': file2_path,
        'original_file1': original_file1_name,
        'original_file2': original_file2_name
    })



@csrf_exempt
@require_POST
@login_required
def process_data(request):
    

    # --- HELPER: Excel Styling ---
    # --- HELPER: Excel Styling ---
    def apply_header_style(workbook, sheet_name, title, blue_color, match_columns=None):
        if sheet_name not in workbook.sheetnames: return
        sheet = workbook[sheet_name]
        
        # Check if title is already applied (simple check to avoid double styling)
        if sheet.cell(row=1, column=1).value == title:
            return

        sheet.insert_rows(1)
        
        # Create Yellow Style
        match_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        match_font = Font(name='Calibri', size=12, color="000000", bold=True)
        
        # Create Standard Blue Style
        blue_fill = PatternFill(start_color=blue_color, end_color=blue_color, fill_type="solid")
        blue_font = Font(name='Calibri', size=12, color="FFFFFF", bold=True)

        match_cols_set = set(match_columns) if match_columns else set()

        # Title Row
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=sheet.max_column or 1)
        title_cell = sheet.cell(row=1, column=1, value=title)
        title_cell.font = Font(name='Calibri', size=18, color="FFFFFF", bold=True)
        title_cell.fill = blue_fill
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Header Row (Row 2)
        for col_num in range(1, (sheet.max_column or 1) + 1):
            cell = sheet.cell(row=2, column=col_num)
            col_name = cell.value

            # --- LOGIC FIX: Check for Yellow Highlight ---
            if col_name in match_cols_set:
                cell.fill = match_fill
                cell.font = match_font
            else:
                cell.fill = blue_fill
                cell.font = blue_font
            
            cell.alignment = Alignment(horizontal='center')

        # Auto-width adjustment
        for col in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
            sheet.column_dimensions[column_letter].width = max_length + 2



    def create_summary_sheet(workbook, metrics):
        if 'Summary' in workbook.sheetnames:
            sheet = workbook['Summary']
        else:
            sheet = workbook.create_sheet('Summary', 0)
        
        blue_color = '000c66'
        labels = [
            "Total File 1 Rows", "Total File 2 Rows", "Total Reconciled Transactions",
            "File 1 Unreconciled Transactions", "File 2 Unreconciled Transactions"
        ]
        values = [
            metrics['total_f1'], metrics['total_f2'], metrics['reconciled'],
            metrics['unreconciled_f1'], metrics['unreconciled_f2']
        ]
        
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(labels))
        title_cell = sheet.cell(row=1, column=1, value="Reconciliation Summary")
        title_cell.font = Font(name='Calibri', size=18, color="FFFFFF", bold=True)
        title_cell.fill = PatternFill(start_color=blue_color, end_color=blue_color, fill_type="solid")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for i, (label, value) in enumerate(zip(labels, values)):
            col_num = i + 1
            label_cell = sheet.cell(row=2, column=col_num, value=label)
            label_cell.font = Font(name='Calibri', size=12, color="FFFFFF", bold=True)
            label_cell.fill = PatternFill(start_color=blue_color, end_color=blue_color, fill_type="solid")
            label_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            label_cell.border = thin_border
            
            value_cell = sheet.cell(row=3, column=col_num, value=value)
            value_cell.font = Font(name='Calibri', size=12, bold=True)
            value_cell.alignment = Alignment(horizontal='center', vertical='center')
            value_cell.border = thin_border
            sheet.column_dimensions[get_column_letter(col_num)].width = 25
        sheet.row_dimensions[2].height = 30

    def saving_and_color_reco_sheet(workbook, sheet_name, title, blue_color, file1_cols, file2_cols, selected_f1=None, selected_f2=None, file1_name="File 1", file2_name="File 2", match_columns=None):
        """
        Applies styling to the reconciliation sheet.
        NEW: Accepts 'match_columns' list. Headers found in this list will be highlighted Yellow.
        """
        if sheet_name not in workbook.sheetnames: return
        sheet = workbook[sheet_name]
        
        # Insert 3 rows to accommodate Title, File Groups, and Column Headers
        sheet.insert_rows(1, amount=3)  

        start_col = 1
        end_col = sheet.max_column
        
        # Highlight style for selected output columns (Light Blue)
        highlight_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        # NEW: Highlight style for Matching Columns (Yellow)
        match_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        selected_f1_set = set(selected_f1) if selected_f1 else set()
        selected_f2_set = set(selected_f2) if selected_f2 else set()
        
        # Create a set for match columns for faster lookup
        match_cols_set = set(match_columns) if match_columns else set()

        # --- ROW 1: Main Title ---
        title_cell = sheet.cell(row=1, column=start_col, value=title)
        title_cell.font = Font(name='Calibri', size=18, color="FFFFFF")
        title_cell.fill = PatternFill(start_color=blue_color, end_color=blue_color, fill_type="solid")
        title_cell.alignment = Alignment(horizontal='center')
        sheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

        columns_to_highlight_indices = []
        
        # Lists to track column indices for File 1 and File 2 for merging in Row 2
        f1_indices = []
        f2_indices = []

        # --- ROW 3: Column Headers (processing based on original headers now at Row 4) ---
        for col_num in range(1, end_col + 1):
            cell = sheet.cell(row=3, column=col_num)
            
            # Read the original header from Row 4 (pushed down by 3 inserts)
            data_header_cell = sheet.cell(row=4, column=col_num)
            col_name = data_header_cell.value
            
            cell.value = col_name
            
            # --- NEW LOGIC: Check for Match Column First ---
            if col_name in match_cols_set:
                # Apply Yellow for Match Columns
                cell.fill = match_fill
                cell.font = Font(name='Calibri', size=12, color="000000", bold=True)
                
                # Still track indices for Group Headers (Row 2)
                if col_name in file1_cols:
                    f1_indices.append(col_num)
                elif col_name in file2_cols:
                    f2_indices.append(col_num)

                # Track for data highlighting (Light Blue body)
                if col_name in selected_f1_set or col_name in selected_f2_set:
                    columns_to_highlight_indices.append(col_num)

            # --- Standard Logic for Non-Match Columns ---
            elif col_name in file1_cols:
                # File 1 Column Styling (Row 3) - Orange
                cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                cell.font = Font(name='Calibri', size=12, color="000000")
                f1_indices.append(col_num) 
                if col_name in selected_f1_set:
                    columns_to_highlight_indices.append(col_num)

            elif col_name in file2_cols:
                # File 2 Column Styling (Row 3) - Green
                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                cell.font = Font(name='Calibri', size=12, color="000000")
                f2_indices.append(col_num)
                if col_name in selected_f2_set:
                    columns_to_highlight_indices.append(col_num)

            else:
                # Other Columns Styling (Row 3) - Grey
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.font = Font(name='Calibri', size=12, color="000000")
            
            cell.alignment = Alignment(horizontal='center')

        # --- ROW 2: File Group Headers (File 1 / File 2) ---
        group_header_fill_file1 = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        group_header_fill_file2 = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        group_header_font = Font(name='Calibri', size=16, color="000000", bold=True) 

        # Merge and Label File 1 Group
        if f1_indices:
            min_f1 = min(f1_indices)
            max_f1 = max(f1_indices)
            f1_cell = sheet.cell(row=2, column=min_f1, value=file1_name)
            f1_cell.fill = group_header_fill_file1
            f1_cell.font = group_header_font
            f1_cell.alignment = Alignment(horizontal='center')
            sheet.merge_cells(start_row=2, start_column=min_f1, end_row=2, end_column=max_f1)

        # Merge and Label File 2 Group
        if f2_indices:
            min_f2 = min(f2_indices)
            max_f2 = max(f2_indices)
            f2_cell = sheet.cell(row=2, column=min_f2, value=file2_name)
            f2_cell.fill = group_header_fill_file2
            f2_cell.font = group_header_font
            f2_cell.alignment = Alignment(horizontal='center')
            sheet.merge_cells(start_row=2, start_column=min_f2, end_row=2, end_column=max_f2)

        # --- Delete the original header row (Row 4) ---
        sheet.delete_rows(4)

        # --- Data Highlighting ---
        if columns_to_highlight_indices:
            # Data now starts at Row 4 (after deletion of old header)
            for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row):
                for col_idx in columns_to_highlight_indices:
                    row[col_idx-1].fill = highlight_fill

        # --- Auto Column Width ---
        for col in range(1, end_col + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            for cell in sheet[column_letter]:
                if cell.value is not None:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
            sheet.column_dimensions[column_letter].width = max_length + 2


    

    def highlight_matched_rows(workbook, sheet_name, matched_indices):
        """
        Highlights complete rows in light green based on the provided indices.
        matched_indices: set or list of original 0-based DataFrame indices.
        """
        if sheet_name not in workbook.sheetnames: return
        sheet = workbook[sheet_name]
        
        # Light Green Color
        fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        
        offset = 3 # Row 1 is header, so data starts at Row 2 (Index 0)
        
        for idx in matched_indices:
            row_num = idx + offset
            if row_num <= sheet.max_row:
                for cell in sheet[row_num]:
                    cell.fill = fill
                    
    # --- SQL GENERATION HELPER ---
    # --- SQL GENERATION HELPER ---
    def build_join_conditions(f1_raw_cols, f2_raw_cols, f1_type_map, f2_type_map, f1_table, f2_table, convert_type='decimal'):
        join_conditions = []
        
        NUMERIC_TYPES = ('INTEGER', 'REAL', 'FLOAT', 'DOUBLE', 'DECIMAL', 'BIGINT', 'NUMERIC', 'HUGEINT', 'SMALLINT', 'TINYINT', 'UBIGINT', 'UINTEGER', 'USMALLINT', 'UTINYINT')

        for i in range(len(f1_raw_cols)):
            raw_col_1 = f1_raw_cols[i]
            raw_col_2 = f2_raw_cols[i]

            f1_col_sql = f'{f1_table}."{raw_col_1}"'
            f2_col_sql = f'{f2_table}."{raw_col_2}"'

            f1_type = f1_type_map.get(raw_col_1, 'TEXT')
            f2_type = f2_type_map.get(raw_col_2, 'TEXT')
            
            # --- FIX STARTS HERE ---
            if f1_type in NUMERIC_TYPES and f2_type in NUMERIC_TYPES:
                if convert_type == 'roundoff':
                    join_conditions.append(f"ROUND(CAST({f1_col_sql} AS DOUBLE), 0) = ROUND(CAST({f2_col_sql} AS DOUBLE), 0)")
                else:
                    join_conditions.append(f"ROUND(CAST({f1_col_sql} AS DOUBLE), 2) = ROUND(CAST({f2_col_sql} AS DOUBLE), 2)")
            
            elif f1_type in ('TEXT', 'VARCHAR') or f2_type in ('TEXT', 'VARCHAR'):
                join_conditions.append(f"TRIM(LOWER(CAST({f1_col_sql} AS TEXT))) = TRIM(LOWER(CAST({f2_col_sql} AS TEXT)))")
            
            elif f1_type == 'DATE' or f2_type == 'DATE':
                join_conditions.append(f"CAST({f1_col_sql} AS DATE) = CAST({f2_col_sql} AS DATE)")
                
            else:
                join_conditions.append(f"CAST({f1_col_sql} AS TEXT) = CAST({f2_col_sql} AS TEXT)")
            # --- FIX ENDS HERE ---

        return join_conditions
    



    # --- SQL GENERATION ---
    # --- SQL GENERATION ---
    def generate_sql_query(f1_raw_cols, f2_raw_cols, f1_type_map, f2_type_map, match_type, convert_type='decimal'):
        f1_table = 'file1_table'
        f2_table = 'file2_table'
        join_conditions = []
        
        # Define numeric types list for checking
        NUMERIC_TYPES = ('INTEGER', 'REAL', 'FLOAT', 'DOUBLE', 'DECIMAL', 'BIGINT', 'NUMERIC', 'HUGEINT', 'SMALLINT', 'TINYINT', 'UBIGINT', 'UINTEGER', 'USMALLINT', 'UTINYINT')

        for i in range(len(f1_raw_cols)):
            raw_col_1 = f1_raw_cols[i]
            raw_col_2 = f2_raw_cols[i]

            f1_col_sql = f'{f1_table}."{raw_col_1}"'
            f2_col_sql = f'{f2_table}."{raw_col_2}"'

            f1_type = f1_type_map.get(raw_col_1, 'TEXT')
            f2_type = f2_type_map.get(raw_col_2, 'TEXT')
            
            # --- FIX STARTS HERE ---
            # 1. Check if BOTH are Numeric (Handles Int vs Float mismatch)
            if f1_type in NUMERIC_TYPES and f2_type in NUMERIC_TYPES:
                if convert_type == 'roundoff':
                    join_conditions.append(f"ROUND(CAST({f1_col_sql} AS DOUBLE), 0) = ROUND(CAST({f2_col_sql} AS DOUBLE), 0)")
                else:
                    join_conditions.append(f"ROUND(CAST({f1_col_sql} AS DOUBLE), 2) = ROUND(CAST({f2_col_sql} AS DOUBLE), 2)")
            
            # 2. Handle Text
            elif f1_type in ('TEXT', 'VARCHAR') or f2_type in ('TEXT', 'VARCHAR'):
                join_conditions.append(f"TRIM(LOWER(CAST({f1_col_sql} AS TEXT))) = TRIM(LOWER(CAST({f2_col_sql} AS TEXT)))")
            
            # 3. Handle Date
            elif f1_type == 'DATE' or f2_type == 'DATE':
                join_conditions.append(f"CAST({f1_col_sql} AS DATE) = CAST({f2_col_sql} AS DATE)")
                
            # 4. Fallback
            else:
                join_conditions.append(f"CAST({f1_col_sql} AS TEXT) = CAST({f2_col_sql} AS TEXT)")
            # --- FIX ENDS HERE ---

        common_query = f"SELECT * FROM {f1_table} INNER JOIN {f2_table} ON {' AND '.join(join_conditions)}"
        
        if match_type == 'onlymatch':
            return common_query

        pk_f1_sql = f'{f1_table}."{f1_raw_cols[0]}"'
        pk_f2_sql = f'{f2_table}."{f2_raw_cols[0]}"'

        f1_not_in_query = f"SELECT {f1_table}.* FROM {f1_table} LEFT JOIN {f2_table} ON {' AND '.join(join_conditions)} WHERE {pk_f2_sql} IS NULL"
        f2_not_in_query = f"SELECT {f2_table}.* FROM {f2_table} LEFT JOIN {f1_table} ON {' AND '.join(join_conditions)} WHERE {pk_f1_sql} IS NULL"

        return common_query, f1_not_in_query, f2_not_in_query


   


    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)

    if 'processedData' not in request.POST:
        return JsonResponse({'status': 'error', 'message': 'Missing processedData field'}, status=400)
        
    processed_data_json = request.POST.get('processedData')
    try:
        payload = json.loads(processed_data_json)
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'Invalid JSON in processedData'}, status=400)
    
    data = payload.get('rules', [])
    file1_path = payload.get('file1_path')
    file2_path = payload.get('file2_path')

    # --- 1.1 GET ORIGINAL FILENAMES ---
    original_file1 = payload.get('original_file1', os.path.basename(file1_path) if file1_path else "File1")
    original_file2 = payload.get('original_file2', os.path.basename(file2_path) if file2_path else "File2")

    original_file1_headername = os.path.splitext(original_file1)[0]
    original_file2_headername = os.path.splitext(original_file2)[0]


    print(original_file1, original_file2)

    if not file1_path or not file2_path:
        return JsonResponse({'status': 'error', 'message': 'Missing file paths. Please re-upload files.'}, status=400)
    
    temp_dir = tempfile.gettempdir()
    
    # 1. READ FILES
    try:
        if file1_path.lower().endswith('.csv'):
            df_file1 = pd.read_csv(file1_path)
        else:
            df_file1 = pd.read_excel(file1_path)
        
        if file2_path.lower().endswith('.csv'):
            df_file2 = pd.read_csv(file2_path)
        else:
            df_file2 = pd.read_excel(file2_path)
            
    except Exception as e:
         return JsonResponse({'status': 'error', 'message': f'Error reading files: {str(e)}'}, status=500)


    df_file1 = clean_dataframe_quotes(df_file1)
    df_file2 = clean_dataframe_quotes(df_file2)
    
    df_file1 = convert_to_date_only(df_file1)
    df_file2 = convert_to_date_only(df_file2)  
    
    # --- ADD TRACKING INDEX FOR HIGHLIGHTING ---
    df_file1['__row_index__'] = df_file1.index
    df_file2['__row_index__'] = df_file2.index

    common_cols = set(df_file1.columns).intersection(df_file2.columns)

    
    # --- 2. LICENSE VALIDATION ---
    user = request.user
    has_unlimited = (user.is_superuser) or (getattr(user, 'process_license_limit', 0) == -1)
    has_unlimited = False
    user_group = user.company_group 
    
    unique_key_str = f"{original_file1}|{original_file2}_{len(df_file1)}_{len(df_file2)}"
    
    existing_history_record = None 
    credits_required_for_run = 0 # --- 6. ADD THIS VARIABLE ---

    if not has_unlimited:
        # --- 7. PASS ROW COUNTS AND UNPACK 3 VARIABLES ---
        is_valid, result, credits_required = check_license_validity(
            user, 
            original_file1, 
            original_file2, 
            len(df_file1), 
            len(df_file2)
        )

        if not is_valid:
            return result
        
        existing_history_record = result
        credits_required_for_run = credits_required # --- 8. STORE IT ---

    

    # --- PROCESSING ---
    conn = None
    generated_files = []
    final_reconciled_count = 0 


    try:
        # 2. SETUP DB
        # Using :memory: to avoid file lock issues on Windows
        conn = duckdb.connect(database=DB_PATH, config={'temp_directory': temp_dir})
        
        # Register DataFrames as virtual tables (efficient and fast)
        conn.register('df_file1', df_file1)
        conn.register('df_file2', df_file2)
        
        # Create Tables
        conn.execute("CREATE OR REPLACE TABLE file1_table AS SELECT * FROM df_file1")
        conn.execute("CREATE OR REPLACE TABLE file2_table AS SELECT * FROM df_file2")
        
 
        file1_cols_list = get_table_columns(conn, 'file1_table')
        file2_cols_list = get_table_columns(conn, 'file2_table')
    

 
        for index, item in enumerate(data):
            selected_option = item.get('selectedOption')
            selected_f1_cols = item.get('file1Data', [])
            selected_f2_cols = item.get('file2Data', [])
            output_cols_requested = item.get('outputColumns', [])
            
            user_sheet_name = item.get('sheetName') or f"Rule_{index+1}"
            safe_filename = secure_filename(f"{user_sheet_name}.xlsx")
            this_output_path = os.path.join(temp_dir, f"Out_{index}_{safe_filename}")
            
            # Filter Logic
            clean_dragged_f1 = set([x for x in selected_f1_cols if x])
            clean_dragged_f2 = set([x for x in selected_f2_cols if x])
            all_dragged = clean_dragged_f1.union(clean_dragged_f2)
            unique_requested = set(output_cols_requested)
            
            if not unique_requested.issubset(all_dragged):
                 final_output_cols = output_cols_requested
            else:
                 final_output_cols = [] # Means export all columns

            count_f1_total = len(df_file1)
            count_f2_total = len(df_file2)
            count_reconciled = 0
            count_unrec_f1 = 0
            count_unrec_f2 = 0



            # Sets to store indices of rows that matched
            matched_f1_ids = set()
            matched_f2_ids = set()

            # --- START OF LOGIC REPLACEMENT ---
            with pd.ExcelWriter(this_output_path, engine='openpyxl') as writer:
                
                # ==========================================
                # OPTION 1: FUZZY MATCH LOGIC
                # ==========================================
                if selected_option in ['namematch', 'creditdebitreco']:
                    # ... [Standard Setup Code matches previous] ...
                    try: threshold = float(item.get('threshold', 45))
                    except: threshold = 45.0

                    file1_columns = item.get('file1Data', [])
                    file2_columns = item.get('file2Data', [])
                    if not file1_columns or not file2_columns: continue

                    f1_fuzzy_col = file1_columns[-1]
                    f2_fuzzy_col = file2_columns[-1]
                    f1_exact_cols = file1_columns[:-1]
                    f2_exact_cols = file2_columns[:-1]
                    
                    # ... [Data Cleaning & Matching Logic is same as before] ...
                    df_f1_clean = df_file1.copy()
                    df_f2_clean = df_file2.copy()
                    df_f1_clean['__clean_fuzzy__'] = df_f1_clean[f1_fuzzy_col].apply(clean_string)
                    df_f2_clean['__clean_fuzzy__'] = df_f2_clean[f2_fuzzy_col].apply(clean_string)
                    
                    matched_results = []
                    unmatched_results = []
                    matched_f2_indices = set()
                    
                    # Blocking setup
                    use_blocking = len(f1_exact_cols) > 0
                    file2_lookup = {}
                    df_f2_clean['__orig_idx__'] = df_f2_clean.index
                    f2_records = df_f2_clean.to_dict('records')
                    
                    if use_blocking:
                        blocking_col_f2 = f2_exact_cols[0]
                        blocking_col_f1 = f1_exact_cols[0]
                        for row in f2_records:
                            key = row[blocking_col_f2]
                            if key not in file2_lookup: file2_lookup[key] = []
                            file2_lookup[key].append(row)
                    else:
                        file2_lookup['__all__'] = f2_records
                    
                    f1_records = df_f1_clean.to_dict('records')
                    
                    for file1_row in f1_records:
                        candidates = file2_lookup.get(file1_row[blocking_col_f1], []) if use_blocking else file2_lookup['__all__']
                        matched = False
                        
                        for f2_row in candidates:
                            f2_idx = f2_row['__row_index__']
                            if f2_idx in matched_f2_indices: continue
                            
                            is_exact_match = True
                            start_check = 1 if use_blocking else 0
                            for i in range(start_check, len(f1_exact_cols)):
                                if file1_row[f1_exact_cols[i]] != f2_row[f2_exact_cols[i]]:
                                    is_exact_match = False
                                    break
                            if not is_exact_match: continue

                            if common_word_match(file1_row['__clean_fuzzy__'], f2_row['__clean_fuzzy__'], threshold):
                                res = file1_row.copy()
                                for c2 in file2_columns: res[c2] = f2_row[c2]
                                for c2_all in df_file2.columns:
                                    if c2_all not in res: res[c2_all] = f2_row[c2_all]
                                matched_results.append(res)
                                matched = True
                                matched_f2_indices.add(f2_idx)
                                matched_f1_ids.add(file1_row['__row_index__'])
                                matched_f2_ids.add(f2_idx)
                                break
                        
                        if not matched: unmatched_results.append(file1_row)
                    
                    if matched_results:
                        df_matched = pd.DataFrame(matched_results)
                    else:
                        # If no matches found, we must manually construct the expected columns
                        # Start with File 1 columns
                        expected_cols = list(df_f1_clean.columns)
                        # Add File 2 columns that were supposed to be joined
                        for c in file2_columns:
                            if c not in expected_cols: expected_cols.append(c)
                        df_matched = pd.DataFrame(columns=expected_cols)

                    # 2. Handle FILE 1 UNRECONCILED (The fix for your issue)
                    if unmatched_results:
                        df_unmatched_f1 = pd.DataFrame(unmatched_results)
                    else:
                        # CRITICAL: If list is empty, create DataFrame using headers from the clean File 1 DF.
                        # This ensures the headers appear in Excel so they can be highlighted.
                        df_unmatched_f1 = pd.DataFrame(columns=df_f1_clean.columns)

                    # 3. Handle FILE 2 UNRECONCILED
                    # .drop() naturally preserves columns even if the result is empty, so this is safe.
                    df_unmatched_f2 = df_f2_clean.drop(index=list(matched_f2_indices))

                    # --- UPDATED ANALYSIS CALLS ---
                    f1_fail_map = {}
                    f2_fail_map = {}

                    if len(df_unmatched_f1) > 0:
                        f1_fail_map = analyze_pandas_mismatch(
                            df_unmatched_f1, df_f2_clean, 
                            f1_exact_cols, f1_fuzzy_col, 
                            f2_exact_cols, f2_fuzzy_col, # Pass Target Cols
                            threshold
                        )
                    
                    if len(df_unmatched_f2) > 0:
                        f2_fail_map = analyze_pandas_mismatch(
                            df_unmatched_f2, df_f1_clean, 
                            f2_exact_cols, f2_fuzzy_col, 
                            f1_exact_cols, f1_fuzzy_col, # Pass Target Cols
                            threshold
                        )

                    # Save Files
                    cols_to_drop = ['__clean_fuzzy__', '__row_index__', '__orig_idx__']
                    df_matched = df_matched.drop(columns=[c for c in cols_to_drop if c in df_matched.columns], errors='ignore')
                    df_matched = filter_dataframe_columns(df_matched, final_output_cols)
                    df_matched.to_excel(writer, sheet_name=user_sheet_name, index=False)
                    
                    count_reconciled = len(df_matched)
                    count_unrec_f1 = len(df_unmatched_f1)
                    count_unrec_f2 = len(df_unmatched_f2)

                    save_excel_with_highlights(writer, df_unmatched_f1, 'File1 Unreconciled', f1_fail_map, final_output_cols)
                    save_excel_with_highlights(writer, df_unmatched_f2, 'File2 Unreconciled', f2_fail_map, final_output_cols)

                # ==========================================
                # OPTION 2: SQL LOGIC
                # ==========================================
                elif selected_option in ['file1reconciled', 'onlymatch', 'diffReconcile']:
                    
                    # ... [Standard Setup Matches Previous] ...
                    file1_cols_raw = item.get('file1Data', [])
                    file2_cols_raw = item.get('file2Data', [])
                    file1_type_map = get_table_column_types(conn, 'file1_table')
                    file2_type_map = get_table_column_types(conn, 'file2_table')
                    convert_type = item.get('convertType', 'decimal')
                    
                    common_query = ""
                    f1_not_in = ""
                    f2_not_in = ""
                    active_join_clause = ""

                    if selected_option == 'diffReconcile':
                        diff_value = item.get('diffValue')
                        diff_range = item.get('diffRange')
                        common_query, f1_not_in, f2_not_in, active_join_clause = generate_difference_query(
                            file1_cols_raw, file2_cols_raw, file1_type_map, file2_type_map, 
                            convert_type, diff_value, diff_range
                        )
                        if final_output_cols: final_output_cols.append("Difference")
                    else:
                        match_type = selected_option
                        if match_type == 'file1reconciled':
                             common_query, f1_not_in, f2_not_in = generate_sql_query(
                                file1_cols_raw, file2_cols_raw, file1_type_map, file2_type_map, match_type, convert_type
                            )
                        else:
                            common_query = generate_sql_query(
                                file1_cols_raw, file2_cols_raw, file1_type_map, file2_type_map, match_type, convert_type
                            )
                        join_conds = build_join_conditions(file1_cols_raw, file2_cols_raw, file1_type_map, file2_type_map, 'file1_table', 'file2_table', convert_type)
                        active_join_clause = ' AND '.join(join_conds)

                    # Execute Matched
                    common_data = conn.execute(common_query).df()
                    
                    if active_join_clause:
                        try:
                            id_query = f"SELECT file1_table.\"__row_index__\" as f1_id, file2_table.\"__row_index__\" as f2_id FROM file1_table INNER JOIN file2_table ON {active_join_clause}"
                            id_df = conn.execute(id_query).df()
                            matched_f1_ids = set(id_df['f1_id'].tolist())
                            matched_f2_ids = set(id_df['f2_id'].tolist())
                        except: pass

                    # Save Matched
                    if '__row_index__' in common_data.columns: common_data = common_data.drop(columns=['__row_index__'])
                    common_data = common_data.loc[:, ~common_data.columns.str.startswith('__row_index__')]
                    common_data = filter_dataframe_columns(common_data, final_output_cols)
                    if "Difference" in common_data.columns:
                        cols = [c for c in common_data.columns if c != "Difference"] + ["Difference"]
                        common_data = common_data[cols]
                    common_data.to_excel(writer, sheet_name=user_sheet_name, index=False)
                    
                    count_reconciled = len(common_data)

                    # Handle Unreconciled
                    if selected_option != 'onlymatch':
                        file1_data = conn.execute(f1_not_in).df().dropna(axis=1, how='all')
                        file2_data = conn.execute(f2_not_in).df().dropna(axis=1, how='all')
                        
                        count_unrec_f1 = len(file1_data)
                        count_unrec_f2 = len(file2_data)

                        # --- UPDATED ANALYSIS CALLS ---
                        f1_fail_map = {}
                        f2_fail_map = {}
                        
                        if count_unrec_f1 > 0:
                            f1_ids = set(file1_data['__row_index__'].tolist())
                            # Pass Both sets of columns
                            f1_fail_map = analyze_sql_mismatch(conn, 'file1_table', 'file2_table', 
                                                               file1_cols_raw, file2_cols_raw, f1_ids)
                        
                        if count_unrec_f2 > 0:
                            f2_ids = set(file2_data['__row_index__'].tolist())
                            # Swap Tables AND Swap Columns for correct mapping
                            f2_fail_map = analyze_sql_mismatch(conn, 'file2_table', 'file1_table', 
                                                               file2_cols_raw, file1_cols_raw, f2_ids)

                        save_excel_with_highlights(writer, file1_data, 'File1 Unreconciled', f1_fail_map, final_output_cols)
                        save_excel_with_highlights(writer, file2_data, 'File2 Unreconciled', f2_fail_map, final_output_cols)
                    else:
                        count_unrec_f1 = count_f1_total - count_reconciled
                        count_unrec_f2 = count_f2_total - count_reconciled

                # --- 3. WRITE ORIGINAL SHEETS WITH GREEN HIGHLIGHTS ---
                df_f1_display = df_file1.drop(columns=['__row_index__'], errors='ignore')
                df_f2_display = df_file2.drop(columns=['__row_index__'], errors='ignore')
                df_f1_display.to_excel(writer, sheet_name='File1 with matched', index=False)
                df_f2_display.to_excel(writer, sheet_name='File2 with matched', index=False)

            # STYLING
            # STYLING
            try:
                workbook = load_workbook(this_output_path)
                metrics = {
                    'total_f1': count_f1_total, 'total_f2': count_f2_total,
                    'reconciled': count_reconciled, 'unreconciled_f1': count_unrec_f1, 'unreconciled_f2': count_unrec_f2
                }
                create_summary_sheet(workbook, metrics)


                match_cols_combined = [c for c in selected_f1_cols if c] + [c for c in selected_f2_cols if c]

                if user_sheet_name in workbook.sheetnames:
                    saving_and_color_reco_sheet(
                        workbook, user_sheet_name, item.get('headerName', 'Reconciliation'), '000c66',
                        file1_cols_list, file2_cols_list,
                        selected_f1=selected_f1_cols, selected_f2=selected_f2_cols,file1_name=original_file1_headername, 
                        file2_name=original_file2_headername,match_columns=match_cols_combined
                    )
                
                extra_sheets = {
                    'File1 Unreconciled': f'File {original_file1_headername} Unreconciled Transactions',
                    'File2 Unreconciled': f'File {original_file2_headername} Unreconciled Transactions',
                    'File1 with matched': f'File {original_file1_headername} with Matched',
                    'File2 with matched': f'File {original_file2_headername} with Matched'
                }
                
                for s_name, title in extra_sheets.items():
                    apply_header_style(workbook, s_name, title, '000c66', match_columns=match_cols_combined)
                
                # --- APPLY GREEN HIGHLIGHTS ---
                highlight_matched_rows(workbook, 'File1 with matched', matched_f1_ids)
                highlight_matched_rows(workbook, 'File2 with matched', matched_f2_ids)

                workbook.save(this_output_path)
                generated_files.append(this_output_path)
            except Exception as e:
                print(f"Styling error on {this_output_path}: {e}")
                generated_files.append(this_output_path)
    
                
    except Exception as e:
        if conn: conn.close()
        import traceback
        traceback.print_exc() 
        return JsonResponse({'status': 'error', 'message': f'Error in rule execution: {str(e)}'}, status=500)

    finally:
        if conn:
            try:
                conn.unregister("df_file1")
                conn.unregister("df_file2")
                conn.close()
            except Exception as e:
                print(f"Error closing DB: {e}")
        try:
            del df_file1
            del df_file2
        except:
            pass
        gc.collect()

    # 4. ZIP & CLEANUP
    zip_buffer = io.BytesIO()
    zip_filename = f'Reconciliation_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip'
    try:
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for file_path in generated_files:
                zf.write(file_path, os.path.basename(file_path))
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'Error creating ZIP: {str(e)}'}, status=500)
    
    zip_buffer.seek(0)

    # --- 5. HISTORY STORAGE ---
    try:
        combined_filename = f"{original_file1} | {original_file2}"
        
        # Only create a NEW record (and deduct credits) if it's not a duplicate
        if not existing_history_record:
            ProcessingHistory.objects.create(
                user=user,
                group=user_group,
                input_filename=combined_filename,
                rows_processed=credits_required_for_run, 
                unique_key=unique_key_str,
                metadata=json.dumps({"reconciled_count": final_reconciled_count})
            )
    except Exception as e:
        print(f"DB Logging Error: {e}")
    
    try:
        for f in generated_files:
            if os.path.exists(f): os.remove(f)
        if os.path.exists(file1_path): os.remove(file1_path)
        if os.path.exists(file2_path): os.remove(file2_path)
    except Exception as e:
        print(f"File Cleanup error (ignoring): {e}")

    response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename={zip_filename}'
    return response











import os
import io
import zipfile
import pandas as pd
import pdfplumber
from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter



@login_required
def pdf_to_excel_view(request):

    if request.method == 'POST' and request.FILES.getlist('pdf_files'):

        user = request.user
        group = getattr(user, 'company_group', None)
        files = request.FILES.getlist('pdf_files')

        # ==========================================
        # 1. PRE-VALIDATION LOGIC
        # ==========================================
        has_unlimited = user.is_superuser 
        license_obj = None

        if not has_unlimited:
            if not group:
                return JsonResponse({'status': 'error', 'message': 'Permission Denied: User does not belong to a Company Group.'}, status=403)

            try:
                license_obj = GroupLicense.objects.get(group=group)

                if license_obj.expiry_date < timezone.now().date():
                    return JsonResponse({'status': 'error', 'message': f'License Expired on {license_obj.expiry_date}. Please contact admin.'}, status=403)

                if not license_obj.is_active:
                    return JsonResponse({'status': 'error', 'message': 'License is currently inactive.'}, status=403)

                # Just do a quick check to see if they are completely out of credits before starting
                credits_used = ProcessingHistory.objects.filter(group=group).aggregate(total=Sum('rows_processed'))['total'] or 0
                if credits_used >= license_obj.row_limit:
                    return JsonResponse({'status': 'error', 'message': 'License Credit Limit Reached. Cannot process more files.'}, status=403)

            except GroupLicense.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'No valid license found for your group.'}, status=403)
            
        # ==========================================
        # 2. PROCESSING LOGIC
        # ==========================================
        zip_buffer = io.BytesIO()
        total_extracted_rows = 0  # <--- TRACK TOTAL ROWS ACROSS ALL PDFs
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for uploaded_file in files:
                try:
                    valid_tables = []
                    with pdfplumber.open(uploaded_file) as pdf:
                        for page in pdf.pages:
                            tables = page.extract_tables()
                            
                            for table in tables:
                                df = pd.DataFrame(table)
                                if df.empty or df.shape[1] <= 5:
                                    continue
                                
                                if df.iloc[0].apply(lambda x: isinstance(x, (int, float))).all():
                                    df.columns = df.iloc[0]
                                    df = df[1:]
                                
                                df.reset_index(drop=True, inplace=True)
                                valid_tables.append(df)

                    if valid_tables:
                        combined_df = pd.concat(valid_tables, ignore_index=True)
                        combined_df.dropna(how='all', inplace=True)
                        combined_df = combined_df.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
                        combined_df.reset_index(drop=True, inplace=True)

                        if not combined_df.empty:
                            header_row = combined_df.iloc[0].tolist()
                            combined_df.columns = header_row
                            combined_df = combined_df[1:]
                            combined_df = combined_df[~combined_df.apply(lambda row: row.tolist() == header_row, axis=1)]
                            combined_df.reset_index(drop=True, inplace=True)
                            
                            # <--- ADD ROWS EXTRACTED FROM THIS PDF TO TOTAL --->
                            total_extracted_rows += len(combined_df)

                        excel_buffer = io.BytesIO()
                        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                            combined_df.to_excel(writer, index=False, sheet_name='Sheet1')
                            
                            workbook = writer.book
                            worksheet = writer.sheets['Sheet1']
                            
                            header_fill = PatternFill(start_color="000c66", end_color="000c66", fill_type="solid")
                            header_font = Font(color="FFFFFF", bold=True)
                            
                            for col_num, cell in enumerate(worksheet[1], 1):
                                cell.fill = header_fill
                                cell.font = header_font
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                            
                            for col in worksheet.columns:
                                max_length = 0
                                column = col[0].column_letter 
                                for cell in col:
                                    try:
                                        if len(str(cell.value)) > max_length:
                                            max_length = len(str(cell.value))
                                    except:
                                        pass
                                adjusted_width = (max_length + 2)
                                worksheet.column_dimensions[column].width = adjusted_width

                        original_name = os.path.splitext(uploaded_file.name)[0]
                        zip_file.writestr(f"{original_name}.xlsx", excel_buffer.getvalue())

                except Exception as e:
                    print(f"Error processing {uploaded_file.name}: {e}")
                    zip_file.writestr(f"ERROR_{uploaded_file.name}.txt", str(e))

        # ==========================================
        # 3. DEDUCT CREDITS AFTER PROCESSING
        # ==========================================
        if not has_unlimited and license_obj:
            # Create a string of all filenames in this batch
            filenames_str = " | ".join([f.name for f in files])
            
            # --- NEW LOGIC: Consistent Key based on File Names and Row Counts ---
            # This makes the PDF tool behave like the Excel tool.
            pdf_unique_key = f"{filenames_str}_{total_extracted_rows}"[:500]

            # CHECK IF THIS EXACT BATCH WAS ALREADY PROCESSED
            existing_record = ProcessingHistory.objects.filter(
                group=group,
                unique_key=pdf_unique_key
            ).first()

            if not existing_record:
                # Only if it's a NEW batch, we calculate and deduct
                is_row_based = getattr(license_obj, 'is_row_based', False)
               

                if is_row_based:
                    ratio = license_obj.rows_per_credit or 500
                    credits_required = math.ceil(total_extracted_rows / float(ratio))
                    if credits_required == 0: credits_required = 1
                else:
                    # File Based = 1 Credit for the entire batch
                    credits_required = 1
                    
                # Verify they have enough remaining credits
                credits_used = ProcessingHistory.objects.filter(group=group).aggregate(total=Sum('rows_processed'))['total'] or 0
                if credits_used + credits_required > license_obj.row_limit:
                    return JsonResponse({
                        'status': 'error', 
                        'message': f'Not enough credits. Extracted {total_extracted_rows} rows. Required: {credits_required}. Available: {license_obj.row_limit - credits_used}.'
                    }, status=403)
                
                # Official Deduction
                ProcessingHistory.objects.create(
                    user=user,
                    group=group,
                    input_filename=f"PDF Batch: {filenames_str}"[:500],
                    rows_processed=credits_required,
                    unique_key=pdf_unique_key # Saves the consistent key
                )
            else:
                print("Same PDF batch detected - Skipping credit deduction.")

        # Finalize ZIP response
        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer, content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename="Processed_PDFs.zip"'
        return response

    return render(request, 'pdf_converter.html')